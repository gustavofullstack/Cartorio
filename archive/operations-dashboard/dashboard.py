#!/usr/bin/env python3
"""Operations Dashboard generator.

Input:  data/operations_data.xlsx (sheets: order_line_export, inventory_snapshot_export,
        fulfillment_export, return_refund_export).
Output: index.html and dashboard_data.json.

The four source sheets are normalized into dated analytical rows (date, orders, gmv,
units, fulfilment, exceptions, returns, inventory) keyed by calendar date so a single
time-range control can drive every panel.

Dashboard plan:
- KPI tiles: 62-day net revenue, orders placed, units shipped, return rate,
  fulfillment on-time (delivered), inventory sellable + inbound pipeline.
- Charts: daily revenue & orders, marketplace & category mix, order-status mix,
  fulfillment performance, return reasons, inventory turnover vs exceptions,
  warehouse stock heatmap, daily return/refund exposure, anomaly markers with
  recommended actions.
- Tables: top SKUs by net contribution, warehouse load + days-on-hand, returns
  heatmap by reason/country, recommended actions backlog.
- Notes: freshness, automation handoff, recommended actions summary.
"""

from __future__ import annotations

import html
import json
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT / "data" / "operations_data.xlsx"
SNAPSHOT_DIR = ROOT / "data" / "snapshots"
ECHARTS_JS = ROOT / "echarts.min.js"
DASHBOARD_RUNTIME_JS = ROOT / "dashboard_runtime.js"
DASHBOARD_HTML = ROOT / "index.html"
DASHBOARD_DATA = ROOT / "dashboard_data.json"

DASHBOARD_TITLE = "Operations Command Center"
DASHBOARD_SUBTITLE = "Sales · Inventory · Orders · Returns — unified operating view"
TIMEZONE_LABEL = "America/Sao_Paulo"
DEFAULT_RANGE = "30D"
SOURCE_LABEL = "operations_data.xlsx (order/fulfillment/inventory/return exports)"
SHEET_ORDER = "order_line_export"
SHEET_FULFILL = "fulfillment_export"
SHEET_INVENTORY = "inventory_snapshot_export"
SHEET_RETURN = "return_refund_export"

EXCEL_EPOCH = datetime(1899, 12, 30)


def _to_date(value) -> str:
    """Return ISO date string YYYY-MM-DD from Excel serial or datetime."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, (int, float)):
        dt = EXCEL_EPOCH + timedelta(days=float(value))
    elif isinstance(value, str) and value:
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return ""
    else:
        return ""
    return f"{dt.year:04d}-{dt.month:02d}-{dt.day:02d}"


def _to_dt(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return EXCEL_EPOCH + timedelta(days=float(value))
    return None


def fmt_int(value: float) -> str:
    return f"{value:,.0f}"


def fmt_money(value: float) -> str:
    return f"${value:,.2f}"


def fmt_money_short(value: float) -> str:
    if abs(value) >= 1_000_000:
        return f"${value/1_000_000:.2f}M"
    if abs(value) >= 1_000:
        return f"${value/1_000:.1f}k"
    return f"${value:.0f}"


def pct(value: float) -> str:
    return f"{value * 100:+.1f}%"


def pct_signed(value: float) -> str:
    sign = "+" if value >= 0 else "-"
    return f"{sign}{abs(value) * 100:.1f}%"


def _coerce_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value != value:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# Data loading + normalization
# ---------------------------------------------------------------------------

def read_sources() -> dict:
    """Load the four sheets from the operations workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    sheets: dict[str, list[dict]] = {}
    for name in (SHEET_ORDER, SHEET_FULFILL, SHEET_INVENTORY, SHEET_RETURN):
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter))
        except StopIteration:
            headers = []
        rows = []
        for row in rows_iter:
            if row is None or all(cell is None for cell in row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})
        sheets[name] = rows
    wb.close()
    return sheets


def _safe_float(value, default=0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value, default=0) -> int:
    try:
        if value is None:
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


ON_TIME_STATUSES = {"delivered"}
IN_FLIGHT_STATUSES = {"in_transit", "out_for_delivery", "pre_transit"}
EXCEPTION_STATUSES = {"exception", "returned_to_sender"}
HEALTHY_STATUSES = {"paid", "fulfilled", "partially_fulfilled"}
AT_RISK_STATUSES = {"cancelled", "refunded", "chargeback_open"}


def normalize_snapshots(sheets: dict) -> list[dict]:
    """Build per-day operating rows from the four source sheets.

    Each output row has shape:
        {date, revenue, discount, shipping, tax, orders, units, marketplaces:set,
         categories:set, order_status:mix, by_marketplace:revenue, by_category:revenue,
         by_status:count, fulfill_total, fulfill_delivered, fulfill_exceptions,
         fulfill_inflight, fulfill_avg_days, ship_cost, returns, refund_amount,
         refund_processed, refund_failed, return_reasons:mix,
         inventory_sellable, inventory_inbound, inventory_unsellable,
         inventory_skus, warehouses:set, anomalies:list[str]}
    """
    payload = sheets if isinstance(sheets.get(SHEET_ORDER), list) else sheets.get("sheets", {})
    if not payload:
        # legacy callers pass a dict that already maps sheet -> rows
        payload = sheets
    orders = payload.get(SHEET_ORDER, []) or []
    fulfillments = payload.get(SHEET_FULFILL, []) or []
    inventories = payload.get(SHEET_INVENTORY, []) or []
    returns = payload.get(SHEET_RETURN, []) or []
    if isinstance(orders, dict):
        orders = orders.get("rows", []) or []
    if isinstance(fulfillments, dict):
        fulfillments = fulfillments.get("rows", []) or []
    if isinstance(inventories, dict):
        inventories = inventories.get("rows", []) or []
    if isinstance(returns, dict):
        returns = returns.get("rows", []) or []  # noqa: F841

    by_date: dict[str, dict] = defaultdict(lambda: {
        "date": "",
        "revenue": 0.0,
        "discount": 0.0,
        "shipping": 0.0,
        "tax": 0.0,
        "orders": set(),
        "lines": 0,
        "units": 0,
        "marketplaces": set(),
        "categories": set(),
        "by_marketplace": defaultdict(float),
        "by_category": defaultdict(float),
        "by_status": defaultdict(int),
        "fulfill_total": 0,
        "fulfill_delivered": 0,
        "fulfill_inflight": 0,
        "fulfill_exceptions": 0,
        "fulfill_days_sum": 0.0,
        "fulfill_days_n": 0,
        "ship_cost": 0.0,
        "ship_cost_delivered": 0.0,
        "returns": 0,
        "refund_amount": 0.0,
        "refund_processed": 0,
        "refund_failed": 0,
        "return_reasons": defaultdict(int),
        "return_countries": defaultdict(int),
        "inventory_sellable": 0,
        "inventory_inbound": 0,
        "inventory_unsellable": 0,
        "inventory_skus": set(),
        "warehouses": set(),
        "snapshots_seen": 0,
    })

    sku_totals: dict[str, dict] = defaultdict(lambda: {
        "sku": "", "category": "", "revenue": 0.0, "units": 0,
        "discounts": 0.0, "returns": 0, "qty_ordered": 0,
    })

    warehouse_load: dict[str, dict] = defaultdict(lambda: {
        "warehouse": "", "sellable": 0, "inbound": 0, "unsellable": 0,
        "snapshots": 0, "delivered": 0, "exceptions": 0,
        "ship_cost": 0.0, "lines": 0,
    })

    return_reason_country: dict[tuple[str, str], int] = defaultdict(int)
    status_counts: dict[str, int] = defaultdict(int)
    marketplace_lines: dict[str, int] = defaultdict(int)
    category_lines: dict[str, int] = defaultdict(int)
    carrier_counts: dict[str, int] = defaultdict(int)
    carrier_delivered: dict[str, int] = defaultdict(int)
    carrier_avg_days: dict[str, list[float]] = defaultdict(list)
    sku_return_count: dict[str, int] = defaultdict(int)

    for row in orders:
        order_id = _coerce_str(row.get("order_id"))
        order_date = _to_date(row.get("order_created_at"))
        if not order_date:
            continue
        cell = by_date[order_date]
        cell["date"] = order_date
        revenue = _safe_float(row.get("line_total"))
        discount = _safe_float(row.get("item_discount"))
        shipping = _safe_float(row.get("shipping_charged"))
        tax = _safe_float(row.get("tax_collected"))
        units = _safe_int(row.get("quantity"))
        marketplace = _coerce_str(row.get("marketplace"))
        category = _coerce_str(row.get("category"))
        sku = _coerce_str(row.get("sku"))
        order_status = _coerce_str(row.get("order_status"))
        if order_id:
            cell["orders"].add(order_id)
        cell["lines"] += 1
        cell["units"] += units
        cell["revenue"] += revenue
        cell["discount"] += discount
        cell["shipping"] += shipping
        cell["tax"] += tax
        if marketplace:
            cell["marketplaces"].add(marketplace)
            cell["by_marketplace"][marketplace] += revenue
            marketplace_lines[marketplace] += 1
        if category:
            cell["categories"].add(category)
            cell["by_category"][category] += revenue
            category_lines[category] += 1
        if order_status:
            cell["by_status"][order_status] += 1
            status_counts[order_status] += 1
        if sku:
            sku_totals[sku]["sku"] = sku
            sku_totals[sku]["category"] = category
            sku_totals[sku]["revenue"] += revenue
            sku_totals[sku]["units"] += units
            sku_totals[sku]["discounts"] += discount
            sku_totals[sku]["qty_ordered"] += units

    for row in fulfillments:
        label_at = _to_dt(row.get("label_created_at"))
        if not label_at:
            continue
        ship_at = _to_dt(row.get("shipped_at"))
        delivered_at = _to_dt(row.get("delivered_at"))
        date_key = _to_date(row.get("label_created_at"))
        if not date_key:
            continue
        cell = by_date[date_key]
        cell["date"] = date_key
        cell["fulfill_total"] += 1
        tracking = _coerce_str(row.get("tracking_status"))
        carrier = _coerce_str(row.get("carrier"))
        warehouse = _coerce_str(row.get("warehouse"))
        cost = _safe_float(row.get("shipping_cost"))
        cell["ship_cost"] += cost
        if tracking == "delivered":
            cell["fulfill_delivered"] += 1
            cell["ship_cost_delivered"] += cost
        elif tracking in EXCEPTION_STATUSES:
            cell["fulfill_exceptions"] += 1
        elif tracking in IN_FLIGHT_STATUSES:
            cell["fulfill_inflight"] += 1
        if delivered_at and ship_at:
            span = (delivered_at - ship_at).total_seconds() / 86400.0
            if span > 0:
                cell["fulfill_days_sum"] += span
                cell["fulfill_days_n"] += 1
                if carrier:
                    carrier_avg_days[carrier].append(span)
        if warehouse:
            cell["warehouses"].add(warehouse)
            bucket = warehouse_load[warehouse]
            bucket["warehouse"] = warehouse
            bucket["snapshots"] += 1
            bucket["lines"] += 1
            bucket["ship_cost"] += cost
            if tracking == "delivered":
                bucket["delivered"] += 1
            elif tracking in EXCEPTION_STATUSES:
                bucket["exceptions"] += 1
        if carrier:
            carrier_counts[carrier] += 1
            if tracking == "delivered":
                carrier_delivered[carrier] += 1

    for row in inventories:
        snap_date = _to_date(row.get("snapshot_time"))
        if not snap_date:
            continue
        cell = by_date[snap_date]
        cell["date"] = snap_date
        cell["snapshots_seen"] += 1
        cell["inventory_sellable"] += _safe_int(row.get("sellable_units"))
        cell["inventory_inbound"] += _safe_int(row.get("inbound_units"))
        cell["inventory_unsellable"] += _safe_int(row.get("unsellable_units"))
        sku = _coerce_str(row.get("sku"))
        warehouse = _coerce_str(row.get("warehouse"))
        if sku:
            cell["inventory_skus"].add(sku)
        if warehouse:
            cell["warehouses"].add(warehouse)
            bucket = warehouse_load[warehouse]
            bucket["warehouse"] = warehouse
            bucket["snapshots"] += 1
            bucket["sellable"] += _safe_int(row.get("sellable_units"))
            bucket["inbound"] += _safe_int(row.get("inbound_units"))
            bucket["unsellable"] += _safe_int(row.get("unsellable_units"))

    for row in returns:
        request_date = _to_date(row.get("request_time"))
        if not request_date:
            continue
        cell = by_date[request_date]
        cell["date"] = request_date
        cell["returns"] += 1
        refund_amount = _safe_float(row.get("refund_amount"))
        cell["refund_amount"] += refund_amount
        refund_status = _coerce_str(row.get("refund_status"))
        if refund_status == "processed":
            cell["refund_processed"] += 1
        elif refund_status == "failed":
            cell["refund_failed"] += 1
        reason = _coerce_str(row.get("reason_code"))
        country = _coerce_str(row.get("country"))
        if reason:
            cell["return_reasons"][reason] += 1
        if country and reason:
            return_reason_country[(reason, country)] += 1
            cell["return_countries"][country] += 1
        sku = _coerce_str(row.get("sku"))
        if sku:
            sku_return_count[sku] += 1

    rows = []
    for date_key, cell in by_date.items():
        if not cell.get("date"):
            continue
        avg_days = (
            cell["fulfill_days_sum"] / cell["fulfill_days_n"]
            if cell["fulfill_days_n"]
            else 0.0
        )
        on_time_pct = (
            cell["fulfill_delivered"] / cell["fulfill_total"]
            if cell["fulfill_total"]
            else 0.0
        )
        orders_count = len(cell["orders"])
        return_rate = (cell["returns"] / orders_count) if orders_count else 0.0
        rows.append(
            {
                "date": cell["date"],
                "revenue": round(cell["revenue"], 2),
                "discount": round(cell["discount"], 2),
                "shipping": round(cell["shipping"], 2),
                "tax": round(cell["tax"], 2),
                "orders": orders_count,
                "lines": cell["lines"],
                "units": cell["units"],
                "marketplace_count": len(cell["marketplaces"]),
                "category_count": len(cell["categories"]),
                "warehouses_active": len(cell["warehouses"]),
                "by_marketplace": dict(cell["by_marketplace"]),
                "by_category": dict(cell["by_category"]),
                "by_status": dict(cell["by_status"]),
                "fulfill_total": cell["fulfill_total"],
                "fulfill_delivered": cell["fulfill_delivered"],
                "fulfill_exceptions": cell["fulfill_exceptions"],
                "fulfill_inflight": cell["fulfill_inflight"],
                "fulfill_on_time_pct": round(on_time_pct, 4),
                "fulfill_avg_days": round(avg_days, 2),
                "ship_cost": round(cell["ship_cost"], 2),
                "ship_cost_delivered": round(cell["ship_cost_delivered"], 2),
                "returns": cell["returns"],
                "return_rate": round(return_rate, 4),
                "refund_amount": round(cell["refund_amount"], 2),
                "refund_processed": cell["refund_processed"],
                "refund_failed": cell["refund_failed"],
                "return_reasons": dict(cell["return_reasons"]),
                "return_countries": dict(cell["return_countries"]),
                "inventory_sellable": cell["inventory_sellable"],
                "inventory_inbound": cell["inventory_inbound"],
                "inventory_unsellable": cell["inventory_unsellable"],
                "inventory_skus": len(cell["inventory_skus"]),
            }
        )

    rows.sort(key=lambda r: r["date"])

    # Side datasets (SKU, warehouse, carrier) — total coverage, not filtered by range.
    sku_table = []
    for sku, info in sku_totals.items():
        sku_returns = sku_return_count.get(sku, 0)
        net = info["revenue"] - info["discounts"]
        sku_table.append({
            "sku": sku,
            "category": info["category"],
            "revenue": round(info["revenue"], 2),
            "net_revenue": round(net, 2),
            "units": info["units"],
            "returns": sku_returns,
            "return_rate_pct": round((sku_returns / info["units"]) * 100, 1) if info["units"] else 0.0,
        })
    sku_table.sort(key=lambda r: r["net_revenue"], reverse=True)

    warehouse_table = []
    for wh, info in warehouse_load.items():
        delivered = info["delivered"]
        exceptions = info["exceptions"]
        total = max(delivered + exceptions, 1)
        warehouse_table.append({
            "warehouse": wh,
            "inventory": info["sellable"],
            "inbound": info["inbound"],
            "unsellable": info["unsellable"],
            "fulfill_lines": info["lines"],
            "delivered": delivered,
            "exceptions": exceptions,
            "exception_rate_pct": round((exceptions / total) * 100, 1),
            "ship_cost": round(info["ship_cost"], 2),
        })
    warehouse_table.sort(key=lambda r: r["fulfill_lines"], reverse=True)

    return_reason_table = sorted(
        (
            {"reason": reason, "country": country, "count": count}
            for (reason, country), count in return_reason_country.items()
        ),
        key=lambda r: r["count"],
        reverse=True,
    )[:60]

    carrier_table = []
    for carrier, total in carrier_counts.items():
        delivered = carrier_delivered.get(carrier, 0)
        avg_days_list = carrier_avg_days.get(carrier, [])
        avg_days = sum(avg_days_list) / len(avg_days_list) if avg_days_list else 0.0
        carrier_table.append({
            "carrier": carrier,
            "lines": total,
            "delivered": delivered,
            "on_time_pct": round((delivered / total) * 100, 1) if total else 0.0,
            "avg_days": round(avg_days, 2),
        })
    carrier_table.sort(key=lambda r: r["lines"], reverse=True)

    return {
        "rows": rows,
        "sku_table": sku_table,
        "warehouse_table": warehouse_table,
        "return_reason_table": return_reason_table,
        "carrier_table": carrier_table,
        "status_totals": dict(status_counts),
        "marketplace_lines": dict(marketplace_lines),
        "category_lines": dict(category_lines),
        "sheet_counts": {
            "orders": len(orders),
            "fulfillments": len(fulfillments),
            "inventory": len(inventories),
            "returns": len(returns),
        },
    }


# ---------------------------------------------------------------------------
# Derived analytics: anomalies + recommended actions
# ---------------------------------------------------------------------------

def detect_anomalies(rows: list[dict], full_payload: dict) -> list[dict]:
    """Identify days with statistically unusual changes vs trailing-window baseline."""
    anomalies = []
    if len(rows) < 7:
        return anomalies
    for idx, row in enumerate(rows):
        window = rows[max(0, idx - 14):idx]
        if len(window) < 5:
            continue
        baseline_rev = sum(r["revenue"] for r in window) / len(window)
        baseline_units = sum(r["units"] for r in window) / len(window)
        baseline_returns = sum(r["returns"] for r in window) / len(window)
        baseline_exceptions = sum(r["fulfill_exceptions"] for r in window) / len(window)
        signals = []
        if baseline_rev >= 50 and row["revenue"] and abs(row["revenue"] - baseline_rev) / max(baseline_rev, 1) > 0.6:
            pct_delta = (row["revenue"] - baseline_rev) / baseline_rev
            signals.append(("revenue_spike" if pct_delta > 0 else "revenue_drop",
                            f"Revenue {pct_signed(pct_delta)} vs 14-day baseline (${baseline_rev:,.0f} avg → ${row['revenue']:,.0f}).",
                            abs(pct_delta)))
        if baseline_units >= 1 and row["units"] and abs(row["units"] - baseline_units) / max(baseline_units, 1) > 0.7:
            pct_delta = (row["units"] - baseline_units) / baseline_units
            signals.append(("units_spike" if pct_delta > 0 else "units_drop",
                            f"Units {pct_signed(pct_delta)} vs baseline ({baseline_units:.1f} avg → {row['units']}).",
                            abs(pct_delta) * 0.85))
        if baseline_returns >= 0.1 and row["returns"] and (row["returns"] - baseline_returns) / max(baseline_returns, 1) > 1.5:
            delta = row["returns"] - baseline_returns
            signals.append(("return_burst",
                            f"Returns jumped to {row['returns']} (+{delta:.1f} vs baseline).",
                            abs(row["returns"] - baseline_returns) / 8))
        if baseline_exceptions >= 0.1 and row["fulfill_exceptions"] and row["fulfill_exceptions"] - baseline_exceptions >= 3:
            signals.append(("exception_burst",
                            f"Fulfillment exceptions {row['fulfill_exceptions']} vs baseline {baseline_exceptions:.1f}.",
                            0.6))
        if not signals:
            continue
        magnitude = max(s for _, _, s in signals)
        anomalies.append({
            "date": row["date"],
            "magnitude": round(magnitude, 2),
            "signals": [{"type": stype, "detail": detail} for stype, detail, _ in signals],
            "orders": row["orders"],
            "revenue": row["revenue"],
            "returns": row["returns"],
        })
    anomalies.sort(key=lambda a: a["magnitude"], reverse=True)
    return anomalies


def build_recommended_actions(rows: list[dict], anomalies: list[dict], full_payload: dict) -> list[dict]:
    """Translate strongest signals into discrete recommended actions."""
    actions = []

    # Anomaly-driven actions (top 3).
    for anomaly in anomalies[:3]:
        severity = "high" if anomaly["magnitude"] >= 1.0 else "medium"
        targets = {s["type"] for s in anomaly["signals"]}
        title_bits = []
        if "revenue_drop" in targets:
            title_bits.append("Investigate revenue drop")
        if "revenue_spike" in targets:
            title_bits.append("Verify demand vs inventory")
        if "return_burst" in targets:
            title_bits.append("Triage return burst")
        if "exception_burst" in targets:
            title_bits.append("Audit fulfillment exceptions")
        if "units_drop" in targets:
            title_bits.append("Diagnose units decline")
        actions.append({
            "priority": severity,
            "title": " · ".join(title_bits) or "Investigate anomaly",
            "context": anomaly["date"],
            "detail": " | ".join(s["detail"] for s in anomaly["signals"]),
        })

    # Status mix alerts.
    status_totals = full_payload.get("status_totals", {})
    total_lines = sum(status_totals.values()) or 1
    chargeback_share = status_totals.get("chargeback_open", 0) / total_lines
    cancel_share = status_totals.get("cancelled", 0) / total_lines
    refund_share = status_totals.get("refunded", 0) / total_lines
    if chargeback_share >= 0.10:
        actions.append({
            "priority": "high",
            "title": "Address chargeback exposure",
            "context": "Aggregate 62-day mix",
            "detail": f"Chargeback-open lines at {chargeback_share*100:.1f}% of order volume — escalate to fraud/SaaS finance review.",
        })
    elif chargeback_share >= 0.06:
        actions.append({
            "priority": "medium",
            "title": "Address chargeback exposure",
            "context": "Aggregate 62-day mix",
            "detail": f"Chargeback-open lines at {chargeback_share*100:.1f}% of order volume — schedule finance review.",
        })
    if cancel_share >= 0.16:
        actions.append({
            "priority": "medium",
            "title": "Reduce order cancellations",
            "context": "Aggregate 62-day mix",
            "detail": f"Cancellations at {cancel_share*100:.1f}% — review checkout, inventory allocation, and SKU availability gates.",
        })
    if refund_share >= 0.14:
        actions.append({
            "priority": "medium",
            "title": "Refund throughput watch",
            "context": "Aggregate 62-day mix",
            "detail": f"Refunded lines at {refund_share*100:.1f}% — confirm refund_status=processed reconciliation between marketplace and finance.",
        })

    # Return reason hotspots.
    reason_table = full_payload.get("return_reason_table", [])
    reason_totals: dict[str, int] = defaultdict(int)
    for entry in reason_table:
        reason_totals[entry["reason"]] += entry["count"]
    top_reason, top_count = max(reason_totals.items(), key=lambda kv: kv[1], default=("", 0))
    if top_count >= 25:
        actions.append({
            "priority": "high",
            "title": f"Top return reason: {top_reason.replace('_',' ')}",
            "context": "Aggregate 62-day returns",
            "detail": f"{top_count} returns in the period — open product QA + PDP copy review for affected SKUs.",
        })

    # Fulfillment carrier issue.
    carrier_table = full_payload.get("carrier_table", [])
    if carrier_table:
        worst = min(carrier_table, key=lambda c: c["on_time_pct"])
        if worst["on_time_pct"] < 70:
            actions.append({
                "priority": "medium",
                "title": f"Re-rate carrier {worst['carrier']}",
                "context": "Carrier on-time delivery",
                "detail": f"On-time delivery {worst['on_time_pct']:.1f}% across {worst['lines']} labels — request SLA review or split volume.",
            })

    # Inventory exposure.
    last_row = rows[-1] if rows else {}
    unsellable_share = 0
    if last_row.get("inventory_sellable"):
        unsellable_share = last_row.get("inventory_unsellable", 0) / max(
            last_row.get("inventory_sellable", 0) + last_row.get("inventory_unsellable", 1), 1
        )
    if unsellable_share >= 0.03:
        actions.append({
            "priority": "medium",
            "title": "Manage unsellable inventory",
            "context": f"Latest snapshot {last_row.get('date','')}",
            "detail": f"Unsellable inventory at {unsellable_share*100:.1f}% — schedule liquidation or warehouse return authorization.",
        })

    # Stable bandwidth: marketing/UTM tracking.
    utm_pulls: dict[str, int] = defaultdict(int)
    for row in full_payload.get("rows", [])[:1]:
        pass
    return actions


# ---------------------------------------------------------------------------
# Payload assembly
# ---------------------------------------------------------------------------

def make_dashboard_payload(sheets: dict) -> dict:
    normalized = normalize_snapshots(sheets)
    rows = normalized["rows"]
    sku_table = normalized["sku_table"]
    warehouse_table = normalized["warehouse_table"]
    return_reason_table = normalized["return_reason_table"]
    carrier_table = normalized["carrier_table"]
    sheet_counts = normalized["sheet_counts"]

    # Inject reduced datasets (avoid passing sets through JSON).
    sheet_counts_payload = dict(sheet_counts)

    full_payload = {
        "rows": rows,
        "sku_table": sku_table,
        "warehouse_table": warehouse_table,
        "return_reason_table": return_reason_table,
        "carrier_table": carrier_table,
        "status_totals": normalized["status_totals"],
        "marketplace_lines": normalized["marketplace_lines"],
        "category_lines": normalized["category_lines"],
    }

    anomalies = detect_anomalies(rows, full_payload)
    recommended_actions = build_recommended_actions(rows, anomalies, full_payload)

    dates = [r["date"] for r in rows]
    latest = dates[-1] if dates else ""
    if len(dates) >= 30:
        window_start = dates[-30]
        prev_start = dates[-60] if len(dates) >= 60 else dates[0]
        prev_end = dates[-31]
    elif dates:
        midpoint = len(dates) // 2
        window_start = dates[midpoint]
        prev_start = dates[0]
        prev_end = dates[midpoint - 1] if midpoint > 0 else latest
    else:
        window_start = ""
        prev_start = ""
        prev_end = latest

    def sum_window(field, start=window_start, end=latest):
        return sum(float(r.get(field) or 0) for r in rows if start <= r["date"] <= end)

    def sum_window_count(field, start=window_start, end=latest):
        return sum(int(r.get(field) or 0) for r in rows if start <= r["date"] <= end)

    rev_30 = sum_window("revenue")
    rev_prev = sum_window("revenue", prev_start, prev_end)
    orders_30 = sum_window_count("orders")
    orders_prev = sum_window_count("orders", prev_start, prev_end)
    units_30 = sum_window("units")
    units_prev = sum_window("units", prev_start, prev_end)
    returns_30 = sum_window_count("returns")
    refunds_30 = sum_window("refund_amount")
    fulfill_total = sum_window_count("fulfill_total")
    fulfill_delivered = sum_window_count("fulfill_delivered")
    fulfill_exceptions = sum_window_count("fulfill_exceptions")
    fulfill_inflight = sum_window_count("fulfill_inflight")
    inventory_sellable = (rows[-1].get("inventory_sellable") if rows else 0) or 0
    inventory_inbound = (rows[-1].get("inventory_inbound") if rows else 0) or 0
    inventory_unsellable = (rows[-1].get("inventory_unsellable") if rows else 0) or 0

    rev_delta = (rev_30 - rev_prev) / rev_prev if rev_prev else 0.0
    orders_delta = (orders_30 - orders_prev) / orders_prev if orders_prev else 0.0
    units_delta = (units_30 - units_prev) / units_prev if units_prev else 0.0
    return_rate = (returns_30 / orders_30) if orders_30 else 0.0
    on_time_pct = (fulfill_delivered / fulfill_total) if fulfill_total else 0.0
    exception_rate = (fulfill_exceptions / fulfill_total) if fulfill_total else 0.0
    inventory_total = inventory_sellable + inventory_unsellable
    unsellable_share = (inventory_unsellable / inventory_total) if inventory_total else 0.0

    latest_captured = (
        f"snapshot_time latest {rows[-1].get('date') if rows else 'n/a'}"
    )

    source_snippets = {
        "revenueTrend": """rows: list[dict] from normalize_snapshots()
daily = sorted rows by date with revenue & orders
filtered = [r for r in daily if start <= r.date <= end]
series = [{date, revenue, orders} for r in filtered]""",
        "marketplaceMix": """group daily['by_marketplace'] across the active range
sum revenue per marketplace, then sort desc
mix = [{marketplace, revenue}] normalized to share %""",
        "categoryMix": """group daily['by_category'] across the active range
sum revenue per category, then sort desc
mix = [{category, revenue}] in active range""",
        "orderStatusMix": """sum daily['by_status'] values across active range
statuses = ['fulfilled','partially_fulfilled','paid','cancelled','refunded','chargeback_open']
chart shows share of line count per status""",
        "fulfillmentPerf": """per day: fulfill_delivered, fulfill_exceptions, fulfill_inflight
on_time = delivered / total; exception_rate = exceptions / total
ship_cost = shipped + delivered cost on each day""",
        "returnReasons": """aggregate return_reasons counts across active range
reasons: missing_item, wrong_size, defective, not_as_described, late_delivery,
changed_mind, damaged_in_transit""",
        "warehouseLoad": """warehouse_load = warehouse + (sellable/inbound/unsellable/delivered/exceptions/ship_cost)
latest snapshot used to anchor inventory; fulfilment volume aggregated across period""",
        "inventoryHealth": """per day: inventory_sellable + inventory_inbound vs inventory_unsellable
utilization = sellable / max(sellable + inbound, 1)
unsellable share = unsellable / (sellable + unsellable)""",
        "dailyReturnCost": """per day: returns + refund_amount
refund_failure_share = refund_failed / max(returns, 1)
return_rate = returns / orders""",
        "anomalyMarkers": """detect_anomalies(): 14-day rolling baseline per day
flags revenue/units swings >= 60%, return bursts >= +50%, exception jumps >= 3
top 5 by magnitude surfaced as recommended actions""",
        "skuLeaderboard": """sku_table grouped by SKU across the entire export
net_revenue = revenue - discounts
return_rate_pct = returns / units (capped via units>0 guard)""",
        "warehouseTable": """warehouse_table aggregates fulfilment volume, ship_cost, exception_rate
over the full export; the date filter affects the embedded inventory scale""",
        "returnMatrix": """return_reason_table: top 60 (reason, country, count)
slice from heatmap-style rows for the matrix panel""",
        "actionsPanel": """build_recommended_actions(): combines anomaly, status mix, return hotspot,
carrier underperformance, inventory unsellable share
priority = high when magnitude >= 1.0 or share >= 10%""",
    }

    analysis_logic = """Analysis logic
- normalize_snapshots() ingests four sheets (order_line_export, fulfillment_export,
  inventory_snapshot_export, return_refund_export) from data/operations_data.xlsx
  and reduces them to per-day {date, revenue, orders, units, fulfill_*,
  return_*, inventory_*, by_marketplace, by_category, by_status, return_reasons}.
- Currency conversion is intentionally not applied: the file records each line in
  its native currency and there is no FX table in the export.
- detect_anomalies() computes a 14-day rolling baseline and flags days whose
  revenue or units deviate > 60%, or whose returns / exceptions jump above the
  baseline. Severity = max signal magnitude.
- build_recommended_actions() combines anomaly signals with status-mix share
  (chargeback, cancel, refund), top return reason hotspots, the worst-performing
  carrier, and unsellable inventory share at the latest snapshot.
- dashboard_runtime.js applies client-side date filtering against the analytical
  date field; tables render from the bounded payload returned here.
"""

    return {
        "title": DASHBOARD_TITLE,
        "subtitle": DASHBOARD_SUBTITLE,
        "timezone": TIMEZONE_LABEL,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "freshness": {
            "latestDataDate": latest,
            "latestCapturedAt": latest_captured,
            "source": SOURCE_LABEL,
            "sheetCounts": sheet_counts_payload,
        },
        "availableDates": dates,
        "defaultRange": DEFAULT_RANGE,
        "kpis": [
            {
                "id": "revenue30",
                "label": "Revenue (last 30D)",
                "value": fmt_money_short(rev_30),
                "delta": pct_signed(rev_delta),
                "detail": f"{orders_30:,} orders · vs prior 30D: {fmt_money_short(rev_prev)}",
            },
            {
                "id": "orders30",
                "label": "Orders placed",
                "value": fmt_int(orders_30),
                "delta": pct_signed(orders_delta),
                "detail": f"{fmt_int(units_30)} units · vs prior 30D: {fmt_int(orders_prev)}",
            },
            {
                "id": "returnRate",
                "label": "Return rate",
                "value": f"{return_rate*100:.1f}%",
                "delta": f"{fmt_int(returns_30)} returns",
                "detail": f"Refund exposure {fmt_money_short(refunds_30)} (native currency)",
            },
            {
                "id": "onTime",
                "label": "Fulfillment on-time",
                "value": f"{on_time_pct*100:.1f}%",
                "delta": f"{fmt_int(fulfill_delivered)} delivered",
                "detail": f"{fmt_int(fulfill_exceptions)} exceptions · {fmt_int(fulfill_inflight)} in-flight",
            },
            {
                "id": "inventory",
                "label": "Sellable inventory",
                "value": fmt_int(inventory_sellable),
                "delta": f"{fmt_int(inventory_inbound)} inbound",
                "detail": f"{fmt_int(inventory_unsellable)} unsellable · {unsellable_share*100:.1f}% unsellable share",
            },
        ],
        "datasets": {
            "daily": rows,
            "sku": sku_table,
            "warehouse": warehouse_table,
            "returnReasons": return_reason_table,
            "carriers": carrier_table,
        },
        "statusTotals": normalized["status_totals"],
        "marketplaceLines": normalized["marketplace_lines"],
        "categoryLines": normalized["category_lines"],
        "anomalies": anomalies[:5],
        "recommendedActions": recommended_actions,
        "sourceSnippets": source_snippets,
        "analysisLogic": analysis_logic,
    }


# ---------------------------------------------------------------------------
# Markup helpers
# ---------------------------------------------------------------------------

def js_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False).replace("</", "<\\/")


def json_script(value) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")


def render_kpi_block(block: dict) -> str:
    delta_class = ""
    delta = block.get("delta", "")
    if delta.startswith("+"):
        delta_class = " kpi-delta-pos"
    elif delta.startswith("-"):
        delta_class = " kpi-delta-neg"
    return f"""
    <section class="kpi-tile" id="{html.escape(block["id"])}">
      <p>{html.escape(block["label"])}</p>
      <strong>{html.escape(block["value"])}</strong>
      <span class="{delta_class.strip()}">{html.escape(delta)}</span>
      <small>{html.escape(block["detail"])}</small>
    </section>
    """


def render_panel_actions(block: dict) -> str:
    edit = ""
    edit_command = ""
    if len(block.get("allowed_types", [])) > 1:
        options = "\n".join(
            f'<option value="{html.escape(kind)}"{" selected" if kind == block.get("initial_type") else ""}>{html.escape(kind)}</option>'
            for kind in block["allowed_types"]
        )
        edit_command = f"""<button onclick="toggleEdit('{html.escape(block["chart_id"])}')">Edit</button>"""
        edit = f"""
        <div class="edit-panel" id="edit-{html.escape(block["chart_id"])}">
          <label for="select-{html.escape(block["chart_id"])}">Type</label>
          <select id="select-{html.escape(block["chart_id"])}" onchange="setChartType('{html.escape(block["chart_id"])}', this.value)">
            {options}
          </select>
        </div>
        """
    return f"""
    <div class="chart-actions">
      {edit}
      <div class="toolbox">
        <button class="tool-button" aria-label="Panel actions" onclick="toggleMenu('{html.escape(block["chart_id"])}')"><span class="dot"></span><span class="dot"></span><span class="dot"></span></button>
        <div class="menu" id="menu-{html.escape(block["chart_id"])}">
          {edit_command}
          <button onclick="viewSource('{html.escape(block["source_key"])}')">View Data Source</button>
        </div>
      </div>
    </div>
    """


def infer_panel_span(block: dict) -> int:
    if block.get("span") is not None:
        span = int(block["span"])
        return span if span in (4, 6, 12) else 6
    if block["kind"] == "table":
        columns = block.get("columns", [])
        has_long_text = any(col.get("long_text") for col in columns)
        return 12 if len(columns) >= 6 or has_long_text else 6
    if block["kind"] == "chart":
        chart_type = str(block.get("initial_type") or "")
        dense_chart = chart_type in {"heatmap", "scatter", "bar"} or block.get("dense")
        many_categories = int(block.get("category_count") or 0) > 8
        return 12 if dense_chart or many_categories else 6
    if block["kind"] == "note":
        return 4 if block.get("compact") else 6
    return 6


def panel_span_attr(block: dict) -> str:
    return f'data-span="{infer_panel_span(block)}"'


def render_chart_block(block: dict) -> str:
    badge = ""
    if block.get("badge"):
        badge = f'<span class="panel-badge">{html.escape(block["badge"])}</span>'
    return f"""
    <section class="dashboard-panel chart-panel" {panel_span_attr(block)} id="{html.escape(block["id"])}">
      <header>
        <div>
          <h2>{html.escape(block["title"])} {badge}</h2>
          <p>{html.escape(block["subtitle"])}</p>
        </div>
        {render_panel_actions(block)}
      </header>
      <div class="chart" id="{html.escape(block["chart_id"])}" role="img" aria-label="{html.escape(block["title"])}"></div>
      <footer>{html.escape(block["unit"])} | {html.escape(block["source_context"])}</footer>
    </section>
    """


def render_table_block(block: dict) -> str:
    columns = block["columns"]
    head = "".join(f"<th>{html.escape(col['label'])}</th>" for col in columns)
    return f"""
    <section class="dashboard-panel table-panel" {panel_span_attr(block)} id="{html.escape(block["id"])}">
      <header>
        <div>
          <h2>{html.escape(block["title"])}</h2>
          <p>{html.escape(block["subtitle"])}</p>
        </div>
        <div class="toolbox">
          <button class="tool-button" aria-label="Panel actions" onclick="toggleMenu('{html.escape(block["source_key"])}')"><span class="dot"></span><span class="dot"></span><span class="dot"></span></button>
          <div class="menu" id="menu-{html.escape(block["source_key"])}">
            <button onclick="viewSource('{html.escape(block["source_key"])}')">View Data Source</button>
          </div>
        </div>
      </header>
      <div class="table-scroll">
        <table id="{html.escape(block["table_id"])}">
          <thead><tr>{head}</tr></thead>
          <tbody></tbody>
        </table>
      </div>
      <footer>{html.escape(block["source_context"])}</footer>
    </section>
    """


def render_note_block(block: dict) -> str:
    return f"""
    <section class="dashboard-note" {panel_span_attr(block)} id="{html.escape(block["id"])}">
      <strong>{html.escape(block["title"])}</strong>
      <span>{html.escape(block["body"])}</span>
    </section>
    """


def render_actions_block(block: dict) -> str:
    items = block.get("items", [])
    rows = ""
    for item in items:
        cls = html.escape(item.get("priority", "medium"))
        rows += (
            f'<li class="action-row action-{cls}">'
            f'<span class="action-pill">{html.escape(item["priority"].upper())}</span>'
            f'<div class="action-body">'
            f'<strong>{html.escape(item["title"])}</strong>'
            f'<small>{html.escape(item["context"])} · {html.escape(item["detail"])}</small>'
            f'</div></li>'
        )
    return f"""
    <section class="dashboard-panel actions-panel" {panel_span_attr(block)} id="{html.escape(block["id"])}">
      <header>
        <div>
          <h2>{html.escape(block["title"])}</h2>
          <p>{html.escape(block["subtitle"])}</p>
        </div>
        <div class="toolbox">
          <button class="tool-button" aria-label="Panel actions" onclick="toggleMenu('{html.escape(block["source_key"])}')"><span class="dot"></span><span class="dot"></span><span class="dot"></span></button>
          <div class="menu" id="menu-{html.escape(block["source_key"])}">
            <button onclick="viewSource('{html.escape(block["source_key"])}')">View Data Source</button>
          </div>
        </div>
      </header>
      <ul class="actions-list">{rows}</ul>
      <footer>{html.escape(block["source_context"])}</footer>
    </section>
    """


# ---------------------------------------------------------------------------
# Block composition
# ---------------------------------------------------------------------------

def build_dashboard_blocks(payload: dict) -> list[dict]:
    blocks: list[dict] = []
    blocks.extend({"kind": "kpi", **kpi} for kpi in payload["kpis"])

    blocks.extend([
        {
            "kind": "chart",
            "id": "panel-revenue-trend",
            "chart_id": "revenueTrend",
            "source_key": "revenueTrend",
            "title": "Revenue & orders",
            "subtitle": "Daily revenue (line) with orders placed (bars) over the active range",
            "unit": "Native currency · orders",
            "source_context": "Source: order_line_export normalized to date grain",
            "allowed_types": ["line", "bar"],
            "initial_type": "line",
        },
        {
            "kind": "chart",
            "id": "panel-marketplace-mix",
            "chart_id": "marketplaceMix",
            "source_key": "marketplaceMix",
            "title": "Marketplace mix",
            "subtitle": "Revenue share per marketplace in the active range",
            "unit": "Native currency",
            "source_context": "Source: order_line_export · by_marketplace aggregate",
            "allowed_types": ["bar", "pie"],
            "initial_type": "pie",
            "dense": True,
        },
        {
            "kind": "chart",
            "id": "panel-category-mix",
            "chart_id": "categoryMix",
            "source_key": "categoryMix",
            "title": "Category revenue mix",
            "subtitle": "Revenue share per product category in the active range",
            "unit": "Native currency",
            "source_context": "Source: order_line_export · by_category aggregate",
            "allowed_types": ["bar", "pie"],
            "initial_type": "bar",
        },
        {
            "kind": "chart",
            "id": "panel-order-status",
            "chart_id": "orderStatusMix",
            "source_key": "orderStatusMix",
            "title": "Order status mix",
            "subtitle": "Share of order lines per order_status in the active range",
            "unit": "% of order lines",
            "source_context": "Source: order_line_export · by_status count",
            "allowed_types": ["bar", "pie"],
            "initial_type": "pie",
        },
        {
            "kind": "chart",
            "id": "panel-fulfillment",
            "chart_id": "fulfillmentPerf",
            "source_key": "fulfillmentPerf",
            "title": "Fulfillment performance",
            "subtitle": "Daily labels, delivered share and exception count",
            "unit": "Lines · delivered · exceptions",
            "source_context": "Source: fulfillment_export by tracking_status",
            "allowed_types": ["line", "bar"],
            "initial_type": "line",
        },
        {
            "kind": "chart",
            "id": "panel-return-reasons",
            "chart_id": "returnReasons",
            "source_key": "returnReasons",
            "title": "Return reason breakdown",
            "subtitle": "Return counts by reason_code in the active range",
            "unit": "Returns",
            "source_context": "Source: return_refund_export · reason_code aggregate",
            "allowed_types": ["bar", "pie"],
            "initial_type": "bar",
        },
        {
            "kind": "chart",
            "id": "panel-warehouse-load",
            "chart_id": "warehouseLoad",
            "source_key": "warehouseLoad",
            "title": "Warehouse load & inventory",
            "subtitle": "Sellable + inbound vs unsellable by 3PL warehouse (latest snapshot)",
            "unit": "Units",
            "source_context": "Source: inventory_snapshot_export + fulfillment_export",
            "allowed_types": ["bar"],
            "initial_type": "bar",
        },
        {
            "kind": "chart",
            "id": "panel-inventory-health",
            "chart_id": "inventoryHealth",
            "source_key": "inventoryHealth",
            "title": "Inventory health & utilization",
            "subtitle": "Sellable vs inbound pipeline (line) and unsellable share (line, secondary)",
            "unit": "Units · %",
            "source_context": "Source: inventory_snapshot_export",
            "allowed_types": ["line"],
            "initial_type": "line",
        },
        {
            "kind": "chart",
            "id": "panel-return-cost",
            "chart_id": "dailyReturnCost",
            "source_key": "dailyReturnCost",
            "title": "Daily return cost & refund failures",
            "subtitle": "Returns per day with refund_amount and failed refunds",
            "unit": "Returns · native currency",
            "source_context": "Source: return_refund_export · refund_status",
            "allowed_types": ["line", "bar"],
            "initial_type": "bar",
        },
        {
            "kind": "chart",
            "id": "panel-anomalies",
            "chart_id": "anomalyMarkers",
            "source_key": "anomalyMarkers",
            "title": "Anomaly markers (recommended actions)",
            "subtitle": "Days flagged by revenue / units / return / exception deviation vs 14-day baseline",
            "unit": "Score (signal magnitude)",
            "source_context": "Source: detect_anomalies() over daily payload",
            "allowed_types": ["scatter", "bar"],
            "initial_type": "scatter",
            "badge": "Recommended actions",
        },
        {
            "kind": "actions",
            "id": "panel-recommended-actions",
            "source_key": "actionsPanel",
            "title": "Recommended actions",
            "subtitle": "Top operational follow-ups derived from anomaly detection, status mix, carrier SLA, and inventory state",
            "unit": "Action backlog",
            "source_context": "Source: build_recommended_actions() aggregate",
            "items": payload["recommendedActions"],
            "span": 12,
        },
        {
            "kind": "table",
            "id": "panel-sku-leaderboard",
            "table_id": "skuLeaderboardTable",
            "source_key": "skuLeaderboard",
            "title": "SKU leaderboard",
            "subtitle": "Top SKUs by net contribution (revenue − discounts) — returns flagging the riskiest lines",
            "source_context": "Source: order_line_export · SKU aggregate",
            "columns": [
                {"field": "sku", "label": "SKU"},
                {"field": "category", "label": "Category"},
                {"field": "units", "label": "Units", "numeric": True},
                {"field": "revenue", "label": "Revenue", "numeric": True},
                {"field": "net_revenue", "label": "Net revenue", "numeric": True},
                {"field": "returns", "label": "Returns", "numeric": True},
                {"field": "return_rate_pct", "label": "Return %", "numeric": True},
            ],
        },
        {
            "kind": "table",
            "id": "panel-warehouse-table",
            "table_id": "warehouseTableTable",
            "source_key": "warehouseTable",
            "title": "Warehouse detail",
            "subtitle": "Volume, on-time delivery and ship cost per 3PL warehouse",
            "source_context": "Source: fulfillment_export + inventory_snapshot_export",
            "columns": [
                {"field": "warehouse", "label": "Warehouse"},
                {"field": "fulfill_lines", "label": "Labels", "numeric": True},
                {"field": "delivered", "label": "Delivered", "numeric": True},
                {"field": "exceptions", "label": "Exceptions", "numeric": True},
                {"field": "exception_rate_pct", "label": "Exception %", "numeric": True},
                {"field": "inventory", "label": "Sellable", "numeric": True},
                {"field": "inbound", "label": "Inbound", "numeric": True},
                {"field": "unsellable", "label": "Unsellable", "numeric": True},
                {"field": "ship_cost", "label": "Ship cost", "numeric": True},
            ],
        },
        {
            "kind": "table",
            "id": "panel-return-matrix",
            "table_id": "returnReasonMatrixTable",
            "source_key": "returnMatrix",
            "title": "Return hotspots (reason × country)",
            "subtitle": "Top combinations from RMA feed — drives refund and PDP copy interventions",
            "source_context": "Source: return_refund_export · reason_code × country",
            "columns": [
                {"field": "reason", "label": "Reason"},
                {"field": "country", "label": "Country"},
                {"field": "count", "label": "Returns", "numeric": True},
            ],
        },
        {
            "kind": "table",
            "id": "panel-carrier-detail",
            "table_id": "carrierDetailTable",
            "source_key": "warehouseTable",
            "title": "Carrier performance",
            "subtitle": "On-time delivery rate and average ship time per carrier",
            "source_context": "Source: fulfillment_export · carrier aggregate",
            "columns": [
                {"field": "carrier", "label": "Carrier"},
                {"field": "lines", "label": "Labels", "numeric": True},
                {"field": "delivered", "label": "Delivered", "numeric": True},
                {"field": "on_time_pct", "label": "On-time %", "numeric": True},
                {"field": "avg_days", "label": "Avg days", "numeric": True},
            ],
        },
        {
            "kind": "note",
            "id": "automation-note",
            "title": "Automation handoff",
            "body": "Schedule a daily job to drop a refreshed operations_data.xlsx under data/, then run python dashboard.py to regenerate index.html and dashboard_data.json.",
        },
        {
            "kind": "note",
            "id": "freshness-note",
            "title": "Data freshness",
            "body": f"Source: {SOURCE_LABEL}. Latest dated snapshot: {payload['freshness']['latestDataDate']} · Generated {payload['generatedAt']} · Currency: native per line (no FX).",
        },
    ])
    return blocks


def render_dashboard_blocks(blocks: list[dict]) -> str:
    kpis = "\n".join(render_kpi_block(b) for b in blocks if b["kind"] == "kpi")
    panels = []
    for block in blocks:
        if block["kind"] == "chart":
            panels.append(render_chart_block(block))
        elif block["kind"] == "table":
            panels.append(render_table_block(block))
        elif block["kind"] == "note":
            panels.append(render_note_block(block))
        elif block["kind"] == "actions":
            panels.append(render_actions_block(block))
    return f"""
    <section class="kpi-grid">{kpis}</section>
    <section class="panel-grid">{"".join(panels)}</section>
    """


# ---------------------------------------------------------------------------
# HTML assembly
# ---------------------------------------------------------------------------

CSS = """
:root {
  color-scheme: light;
  --ink: #2f3437;
  --muted: #68707a;
  --faint: #8b95a3;
  --line: #e1e5ea;
  --line-strong: #cbd3dc;
  --panel: #FAFAFA;
  --page: #ffffff;
  --surface: #FAFAFA;
  --soft: #f2f3f5;
  --soft-blue: #f3f6fb;
  --control-bg: rgba(255, 255, 255, 0.94);
  --topbar-bg: rgba(255, 255, 255, 0.96);
  --menu-bg: #ffffff;
  --modal-bg: #ffffff;
  --modal-backdrop: rgba(55, 53, 47, 0.34);
  --table-head: #FAFAFA;
  --table-hover: #f1f2ff;
  --chart-bg: #FAFAFA;
  --chart-text: #2f3437;
  --chart-muted: #68707a;
  --chart-line: #e1e5ea;
  --chart-primary: #2F6BFF;
  --chart-secondary: #00BFA6;
  --chart-tertiary: #FF7A3D;
  --chart-quaternary: #F45BB3;
  --chart-1: #F45BB3;
  --chart-2: #2F6BFF;
  --chart-3: #00BFA6;
  --chart-4: #FF7A3D;
  --chart-5: #9BD82E;
  --chart-6: #7C3AED;
  --chart-7: #FFD23F;
  --brand: #6979F8;
  --brand-hover: #9EA9FF;
  --brand-end: #CDD2FD;
  --brand-text: #ffffff;
  --accent: #2F6BFF;
  --accent-2: #00BFA6;
  --warn: #b7791f;
  --danger: #c2410c;
  --good: #137757;
}
html[data-theme="trae-dark"] {
  color-scheme: dark;
  --ink: #f5f9fe;
  --muted: #9599a6;
  --faint: #666b75;
  --line: #2a2d31;
  --line-strong: #3a3f45;
  --panel: #1a1b1d;
  --page: #0c0c0d;
  --surface: #222427;
  --soft: #2a2d31;
  --soft-blue: #202123;
  --control-bg: #202123;
  --topbar-bg: rgba(12, 12, 13, 0.92);
  --menu-bg: #202123;
  --modal-bg: #1a1b1d;
  --modal-backdrop: rgba(0, 0, 0, 0.58);
  --table-head: #222427;
  --table-hover: #202123;
  --chart-bg: #222427;
  --chart-text: #d1d3db;
  --chart-muted: #9599a6;
  --chart-line: #2a2d31;
  --chart-primary: #28d9ff;
  --chart-secondary: #32f08c;
  --chart-tertiary: #f6c85f;
  --chart-quaternary: #ff6b9a;
  --chart-1: #32f08c;
  --chart-2: #28d9ff;
  --chart-3: #a78bfa;
  --chart-4: #f6c85f;
  --chart-5: #ff6b9a;
  --chart-6: #6ea8ff;
  --chart-7: #d1d3db;
  --brand: #32f08c;
  --brand-hover: #0fdc78;
  --brand-end: #32f08c;
  --brand-text: #0c0c0d;
  --accent: #32f08c;
  --accent-2: #0fdc78;
  --warn: #f6c85f;
  --danger: #ff6b9a;
  --good: #32f08c;
}
* { box-sizing: border-box; }
html { background: var(--page); }
body {
  margin: 0;
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Helvetica Neue", Arial, sans-serif;
  background: var(--page);
  color: var(--ink);
  font-size: 1rem;
  line-height: 1.55;
  -webkit-font-smoothing: antialiased;
  text-rendering: optimizeLegibility;
}
.topbar {
  position: sticky;
  top: 0;
  z-index: 20;
  border-bottom: 1px solid var(--line);
  background: var(--topbar-bg);
  backdrop-filter: blur(12px);
}
.topbar-inner {
  max-width: 1320px;
  margin: 0 auto;
  padding: 14px 22px;
  display: grid;
  grid-template-columns: minmax(260px, 1fr) auto;
  gap: 18px;
  align-items: center;
}
h1, h2, h3, p { margin: 0; }
h1 { font-size: 22px; font-weight: 500; letter-spacing: 0; }
.subtitle, .freshness, .range-label, .dashboard-panel p, footer, small {
  color: var(--muted);
  font-size: 13px;
  line-height: 1.45;
  font-weight: 400;
}
.controls {
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  justify-content: flex-end;
  gap: 10px;
}
.range-label { display: none; }
.segmented {
  display: inline-flex;
  border: 1px solid var(--line);
  border-radius: 8px;
  overflow: hidden;
  background: var(--control-bg);
}
.segmented button, .menu button, .edit-panel button {
  border: 0;
  background: transparent;
  color: var(--ink);
  font: inherit;
  cursor: pointer;
}
.segmented button {
  min-width: 44px;
  height: 34px;
  padding: 0 10px;
  border-right: 1px solid var(--line);
  font-size: 13px;
  font-weight: 400;
}
.segmented button:last-child { border-right: 0; }
.segmented button.active { background: var(--brand); color: var(--brand-text); font-weight: 500; }
.theme-switch button.active { background: var(--brand); color: var(--brand-text); }
.theme-switch button {
  width: 38px;
  min-width: 38px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  padding: 0;
}
.theme-switch svg {
  width: 16px;
  height: 16px;
  stroke-width: 2;
}
.date-fields { display: inline-flex; align-items: center; gap: 6px; }
input[type="date"] {
  height: 34px;
  border: 1px solid var(--line);
  border-radius: 8px;
  padding: 0 8px;
  background: var(--control-bg);
  color: var(--ink);
  font: inherit;
  font-size: 13px;
  font-weight: 400;
}
.dashboard-shell {
  max-width: 1320px;
  margin: 0 auto;
  padding: 18px 22px 44px;
}
.kpi-grid {
  display: grid;
  grid-template-columns: repeat(5, minmax(0, 1fr));
  gap: 12px;
  margin-bottom: 18px;
}
.kpi-tile {
  min-height: 126px;
  padding: 15px;
  display: grid;
  align-content: space-between;
  gap: 8px;
  background: var(--surface);
  border: 1px solid var(--line);
  border-radius: 12px;
}
.kpi-tile:first-child {
  background: linear-gradient(135deg, var(--brand) 0%, var(--brand-hover) 58%, var(--brand-end) 100%);
  border-color: var(--brand);
}
.kpi-tile p { color: var(--muted); font-size: 13px; font-weight: 500; }
.kpi-tile strong { font-size: 26px; font-weight: 500; letter-spacing: 0; }
.kpi-tile span { color: var(--ink); font-size: 14px; font-weight: 500; line-height: 1.35; }
.kpi-tile small { font-size: 12px; font-weight: 400; }
.kpi-tile:first-child p,
.kpi-tile:first-child strong,
.kpi-tile:first-child span,
.kpi-tile:first-child small { color: var(--brand-text); }
.kpi-delta-pos { color: var(--good) !important; }
.kpi-delta-neg { color: var(--danger) !important; }
.kpi-tile:first-child .kpi-delta-pos,
.kpi-tile:first-child .kpi-delta-neg { color: var(--brand-text) !important; }
.panel-grid {
  display: grid;
  grid-template-columns: repeat(12, minmax(0, 1fr));
  gap: 20px 16px;
}
.dashboard-panel {
  min-height: 360px;
  display: flex;
  flex-direction: column;
  gap: 10px;
  background: transparent;
  border: 0;
  border-radius: 0;
  padding: 0;
}
[data-span="4"] { grid-column: span 4; }
[data-span="6"] { grid-column: span 6; }
[data-span="12"] { grid-column: 1 / -1; }
.dashboard-note {
  min-height: 180px;
  padding: 16px;
  display: flex;
  flex-direction: column;
  gap: 8px;
  border: 1px solid var(--line);
  border-radius: 16px;
  background: transparent;
}
.dashboard-panel header {
  display: flex;
  align-items: flex-start;
  justify-content: space-between;
  gap: 12px;
  min-height: 42px;
}
.dashboard-panel h2 { font-size: 17px; font-weight: 500; letter-spacing: 0; }
.panel-badge {
  margin-left: 8px;
  padding: 2px 8px;
  border-radius: 999px;
  background: var(--soft-blue);
  color: var(--brand);
  font-size: 11px;
  font-weight: 500;
  letter-spacing: 0.2px;
  vertical-align: middle;
}
.chart {
  width: 100%;
  height: 296px;
  min-height: 296px;
  padding: 8px 0 6px;
  border: 1px solid var(--line);
  border-radius: 16px;
  background: var(--chart-bg);
}
.chart-actions {
  display: flex;
  align-items: center;
  justify-content: flex-end;
  gap: 8px;
  flex: 0 0 auto;
  position: relative;
  z-index: 12;
}
.toolbox { position: relative; flex: 0 0 auto; }
.tool-button {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  gap: 3px;
  width: 34px;
  height: 30px;
  border: 1px solid var(--line-strong);
  border-radius: 999px;
  background: var(--control-bg);
  color: var(--muted);
  cursor: pointer;
  font-size: 0;
  line-height: 0;
  padding: 0;
  opacity: 0;
  transition: opacity 140ms ease, background-color 140ms ease;
}
.tool-button .dot {
  display: block;
  width: 3px;
  height: 3px;
  border-radius: 50%;
  background: currentColor;
}
.dashboard-panel:hover .tool-button,
.dashboard-panel:focus-within .tool-button,
.dashboard-note:hover .tool-button,
.dashboard-note:focus-within .tool-button { opacity: 1; }
.menu {
  display: none;
  position: absolute;
  right: 0;
  top: 34px;
  z-index: 40;
  width: 188px;
  padding: 6px;
  border: 1px solid var(--line-strong);
  border-radius: 8px;
  background: var(--menu-bg);
  box-shadow: 0 8px 18px rgba(32, 33, 36, 0.12);
}
.menu.open { display: block; }
.edit-panel {
  display: none;
  align-items: center;
  gap: 6px;
  height: 28px;
  border: 1px solid var(--line);
  border-radius: 6px;
  background: var(--control-bg);
  color: var(--muted);
  font-size: 13px;
}
.edit-panel.open { display: flex; }
.edit-panel label {
  padding-left: 8px;
  white-space: nowrap;
}
.edit-panel select {
  height: 26px;
  border: 0;
  border-left: 1px solid var(--line);
  border-radius: 0 5px 5px 0;
  padding: 0 24px 0 7px;
  background: transparent;
  color: var(--ink);
  font: inherit;
  font-size: 13px;
  outline: none;
  cursor: pointer;
}
.menu button {
  display: block;
  width: 100%;
  border: 0;
  background: transparent;
  padding: 8px 10px;
  border-radius: 6px;
  text-align: left;
  cursor: pointer;
  color: var(--ink);
  font: inherit;
  font-size: 13px;
}
.menu button:hover, .menu button:focus-visible {
  background: var(--soft-blue);
  outline: none;
}
.menu button:focus-visible {
  outline: 2px solid var(--ink);
}
.table-scroll {
  overflow: auto;
  border: 1px solid var(--line);
  border-radius: 16px;
  background: var(--surface);
}
table { width: 100%; border-collapse: collapse; font-size: 13px; }
th, td { padding: 9px 10px; border-bottom: 1px solid var(--line); text-align: left; white-space: nowrap; }
th { color: var(--muted); font-weight: 500; background: var(--table-head); }
td.num { text-align: right; font-variant-numeric: tabular-nums; }
tbody tr:last-child td { border-bottom: 0; }
tbody tr:hover td { background: var(--table-hover); }
.actions-list {
  list-style: none;
  margin: 0;
  padding: 0;
  display: flex;
  flex-direction: column;
  gap: 8px;
}
.action-row {
  display: flex;
  align-items: flex-start;
  gap: 12px;
  padding: 12px 14px;
  border: 1px solid var(--line);
  border-radius: 12px;
  background: var(--surface);
}
.action-row.action-high {
  border-color: var(--danger);
  background: var(--soft);
}
.action-row.action-medium {
  border-color: var(--warn);
}
.action-pill {
  display: inline-block;
  padding: 4px 10px;
  border-radius: 999px;
  font-size: 11px;
  font-weight: 600;
  letter-spacing: 0.4px;
  text-transform: uppercase;
  background: var(--soft-blue);
  color: var(--brand);
  white-space: nowrap;
}
.action-row.action-high .action-pill {
  background: var(--danger);
  color: #ffffff;
}
.action-row.action-medium .action-pill {
  background: var(--warn);
  color: #ffffff;
}
.action-body { display: flex; flex-direction: column; gap: 4px; min-width: 0; }
.action-body strong { font-size: 14px; font-weight: 500; color: var(--ink); line-height: 1.4; }
.action-body small { color: var(--muted); font-size: 12.5px; line-height: 1.5; }
.modal-backdrop {
  position: fixed;
  inset: 0;
  display: none;
  align-items: center;
  justify-content: center;
  padding: 24px;
  background: var(--modal-backdrop);
  z-index: 50;
}
.modal-backdrop.open { display: flex; }
.modal {
  width: min(860px, 100%);
  max-height: min(780px, 92vh);
  overflow: auto;
  border-radius: 16px;
  background: var(--modal-bg);
  border: 1px solid var(--line-strong);
  box-shadow: 0 18px 48px rgba(55, 53, 47, 0.18);
}
.modal-head {
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
  gap: 16px;
  padding: 18px 20px 14px;
  border-bottom: 1px solid var(--line);
}
.modal-head h3 {
  margin: 0;
  font-size: 16px;
  line-height: 1.4;
  font-weight: 600;
}
.modal-subtitle {
  margin: 4px 0 0;
  color: var(--muted);
  font-size: 14px;
  line-height: 1.45;
}
.modal-body { padding: 18px 20px 20px; }
.source-section + .source-section { margin-top: 16px; }
.source-section h4 {
  margin: 0 0 8px;
  color: var(--ink);
  font-size: 14px;
  line-height: 1.4;
  font-weight: 600;
}
.code-wrap { position: relative; }
pre {
  margin: 0;
  padding: 14px;
  overflow: auto;
  border-radius: 8px;
  border: 1px solid var(--line);
  background: var(--soft);
  color: var(--ink);
  font-size: 12px;
  line-height: 1.5;
}
.close {
  width: 32px;
  height: 32px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  border: 0;
  border-radius: 999px;
  background: transparent;
  color: var(--muted);
  cursor: pointer;
  padding: 0;
}
.close svg { width: 18px; height: 18px; stroke-width: 2.1; }
.close:hover, .close:focus-visible { background: var(--soft); outline: none; }
.close:focus-visible { outline: 2px solid var(--ink); }
.copy-button {
  position: absolute;
  right: 8px;
  top: 8px;
  width: 28px;
  height: 28px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  border: 1px solid var(--line);
  border-radius: 6px;
  background: var(--control-bg);
  color: var(--muted);
  cursor: pointer;
}
.copy-button svg { width: 15px; height: 15px; stroke-width: 2; }
.copy-button:hover, .copy-button:focus-visible {
  background: var(--soft);
  color: var(--ink);
  outline: none;
}
.copy-button:focus-visible {
  outline: 2px solid var(--ink);
}
@media (max-width: 1100px) {
  .kpi-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
}
@media (max-width: 900px) {
  .topbar-inner { grid-template-columns: 1fr; }
  .controls { justify-content: flex-start; }
  .kpi-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
  .dashboard-panel, .dashboard-note, [data-span] { grid-column: 1 / -1; }
}
@media (max-width: 620px) {
  .topbar-inner, .dashboard-shell { padding-left: 14px; padding-right: 14px; }
  .kpi-grid { grid-template-columns: 1fr; }
  .segmented { width: 100%; }
  .segmented button { flex: 1; min-width: 0; }
  .date-fields { width: 100%; }
  input[type="date"] { min-width: 0; width: 100%; }
  .chart { height: 240px; min-height: 240px; }
}
"""


def build_html(payload: dict) -> str:
    echarts = ECHARTS_JS.read_text(encoding="utf-8")
    runtime = DASHBOARD_RUNTIME_JS.read_text(encoding="utf-8")
    blocks = build_dashboard_blocks(payload)
    content = render_dashboard_blocks(blocks)
    initial_charts = [
        {"id": b["chart_id"], "type": b["initial_type"]}
        for b in blocks
        if b["kind"] == "chart"
    ]
    table_config = {
        "skuLeaderboardTable": {
            "dataset": "sku",
            "sortField": "net_revenue",
            "sortDirection": "desc",
            "limit": 12,
            "columns": [
                {"field": "sku"},
                {"field": "category"},
                {"field": "units", "numeric": True},
                {"field": "revenue", "numeric": True},
                {"field": "net_revenue", "numeric": True},
                {"field": "returns", "numeric": True},
                {"field": "return_rate_pct", "numeric": True},
            ],
        },
        "warehouseTableTable": {
            "dataset": "warehouse",
            "sortField": "fulfill_lines",
            "sortDirection": "desc",
            "limit": 12,
            "columns": [
                {"field": "warehouse"},
                {"field": "fulfill_lines", "numeric": True},
                {"field": "delivered", "numeric": True},
                {"field": "exceptions", "numeric": True},
                {"field": "exception_rate_pct", "numeric": True},
                {"field": "inventory", "numeric": True},
                {"field": "inbound", "numeric": True},
                {"field": "unsellable", "numeric": True},
                {"field": "ship_cost", "numeric": True},
            ],
        },
        "returnReasonMatrixTable": {
            "dataset": "returnReasons",
            "sortField": "count",
            "sortDirection": "desc",
            "limit": 14,
            "columns": [
                {"field": "reason"},
                {"field": "country"},
                {"field": "count", "numeric": True},
            ],
        },
        "carrierDetailTable": {
            "dataset": "carriers",
            "sortField": "lines",
            "sortDirection": "desc",
            "limit": 12,
            "columns": [
                {"field": "carrier"},
                {"field": "lines", "numeric": True},
                {"field": "delivered", "numeric": True},
                {"field": "on_time_pct", "numeric": True},
                {"field": "avg_days", "numeric": True},
            ],
        },
    }
    source_map = payload["sourceSnippets"]
    analysis_logic = payload["analysisLogic"]

    chart_js = f"""
    const dashboardPayload = {json_script(payload)};
    function cssToken(name) {{
      return getComputedStyle(document.documentElement).getPropertyValue(name).trim();
    }}
    function chartTheme() {{
      return {{
        text: cssToken("--chart-text"),
        muted: cssToken("--chart-muted"),
        line: cssToken("--chart-line"),
        primary: cssToken("--chart-primary"),
        secondary: cssToken("--chart-secondary"),
        tertiary: cssToken("--chart-tertiary"),
        quaternary: cssToken("--chart-quaternary"),
        palette: [1, 2, 3, 4, 5, 6, 7].map(index => cssToken("--chart-" + index))
      }};
    }}
    function axisStyle(extra) {{
      const theme = chartTheme();
      const base = {{
        axisLabel: {{ color: theme.muted }},
        axisLine: {{ lineStyle: {{ color: theme.line }} }},
        axisTick: {{ lineStyle: {{ color: theme.line }} }},
        splitLine: {{ lineStyle: {{ color: theme.line }} }}
      }};
      const merged = Object.assign({{}}, base, extra || {{}});
      merged.axisLabel = Object.assign({{}}, base.axisLabel, (extra || {{}}).axisLabel || {{}});
      return merged;
    }}
    function chartBase(...colorKeys) {{
      const theme = chartTheme();
      return {{
        textStyle: {{ color: theme.text }},
        color: colorKeys.map(key => theme[key] || key)
      }};
    }}
    function ordinalColor(values, index) {{
      const theme = chartTheme();
      if (!values.length) return theme.primary;
      const palette = theme.palette;
      return palette[index % palette.length];
    }}
    const categoryColorKeys = {{
      "Amazon US": "primary",
      "Walmart Marketplace": "secondary",
      "Amazon UK": "tertiary",
      "TikTok Shop US": "quaternary",
      "Shopify DTC": "chart-5",
      "missing_item": "tertiary",
      "wrong_size": "primary",
      "defective": "danger",
      "not_as_described": "secondary",
      "late_delivery": "quaternary",
      "changed_mind": "chart-5",
      "damaged_in_transit": "danger"
    }};
    function categoricalColor(name, index) {{
      const key = String(name || "").trim();
      const theme = chartTheme();
      const token = categoryColorKeys[key];
      if (token) return theme[token] || token;
      return theme.palette[index % theme.palette.length];
    }}
    function aggregateBy(rows, field) {{
      const totals = new Map();
      rows.forEach(row => {{
        const value = row[field] || {{}};
        Object.entries(value || {{}}).forEach(([key, val]) => {{
          totals.set(key, (totals.get(key) || 0) + Number(val || 0));
        }});
      }});
      return Array.from(totals, ([key, val]) => [key, val]).sort((a, b) => b[1] - a[1]);
    }}
    const chartFactories = {{
      revenueTrend: function(type, filteredRows) {{
        const rows = filteredRows("daily");
        return {{
          ...chartBase("primary", "tertiary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0 }},
          grid: {{ left: 56, right: 56, top: 30, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.date), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: [
            axisStyle({{ type: "value", name: "Revenue", nameTextStyle: {{ color: chartTheme().muted }}, axisLabel: {{ formatter: value => "$" + Math.round(value / 1000) + "k" }} }}),
            axisStyle({{ type: "value", name: "Orders", nameTextStyle: {{ color: chartTheme().muted }}, position: "right", splitLine: {{ show: false }} }})
          ],
          series: [
            {{
              name: "Revenue",
              type: type,
              smooth: type === "line",
              data: rows.map(row => row.revenue),
              areaStyle: type === "line" ? {{ opacity: 0.08 }} : undefined,
              itemStyle: {{ color: chartTheme().primary }},
              yAxisIndex: 0
            }},
            {{
              name: "Orders",
              type: "bar",
              data: rows.map(row => row.orders),
              itemStyle: {{ color: chartTheme().tertiary, opacity: 0.55 }},
              barMaxWidth: 8,
              yAxisIndex: 1
            }}
          ]
        }};
      }},
      marketplaceMix: function(type, filteredRows) {{
        const totals = aggregateBy(filteredRows("daily"), "by_marketplace");
        if (type === "pie") {{
          return {{
            ...chartBase(),
            color: totals.map((row, index) => categoricalColor(row[0], index)),
            tooltip: {{ trigger: "item", valueFormatter: value => "$" + Number(value).toFixed(0) }},
            legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0, type: "scroll" }},
            series: [{{ type: "pie", radius: ["40%", "70%"], data: totals.map(row => ({{ name: row[0], value: row[1] }})) }}]
          }};
        }}
        return {{
          ...chartBase("primary"),
          tooltip: {{ trigger: "axis", valueFormatter: value => "$" + Number(value).toFixed(0) }},
          grid: {{ left: 70, right: 18, top: 30, bottom: 30 }},
          xAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => "$" + Math.round(value / 1000) + "k" }} }}),
          yAxis: axisStyle({{ type: "category", data: totals.map(row => row[0]) }}),
          series: [{{ type: "bar", data: totals.map(row => row[1]), itemStyle: {{ color: (params) => categoricalColor(totals[params.dataIndex][0], params.dataIndex) }}, barMaxWidth: 30 }}]
        }};
      }},
      categoryMix: function(type, filteredRows) {{
        const totals = aggregateBy(filteredRows("daily"), "by_category");
        if (type === "pie") {{
          return {{
            ...chartBase(),
            color: totals.map((row, index) => ordinalColor(totals, index)),
            tooltip: {{ trigger: "item", valueFormatter: value => "$" + Number(value).toFixed(0) }},
            legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0, type: "scroll" }},
            series: [{{ type: "pie", radius: ["40%", "70%"], data: totals.map(row => ({{ name: row[0], value: row[1] }})) }}]
          }};
        }}
        return {{
          ...chartBase("primary"),
          tooltip: {{ trigger: "axis", valueFormatter: value => "$" + Number(value).toFixed(0) }},
          grid: {{ left: 60, right: 18, top: 30, bottom: 30 }},
          xAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => "$" + Math.round(value / 1000) + "k" }} }}),
          yAxis: axisStyle({{ type: "category", data: totals.map(row => row[0]) }}),
          series: [{{ type: "bar", data: totals.map(row => row[1]), itemStyle: {{ color: (params) => ordinalColor(totals, params.dataIndex) }}, barMaxWidth: 30 }}]
        }};
      }},
      orderStatusMix: function(type, filteredRows) {{
        const totals = aggregateBy(filteredRows("daily"), "by_status");
        if (type === "pie") {{
          return {{
            ...chartBase(),
            color: totals.map((row, index) => ordinalColor(totals, index)),
            tooltip: {{ trigger: "item", valueFormatter: value => Number(value).toFixed(0) + " lines" }},
            legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0, type: "scroll" }},
            series: [{{ type: "pie", radius: ["40%", "70%"], data: totals.map(row => ({{ name: row[0], value: row[1] }})) }}]
          }};
        }}
        return {{
          ...chartBase("primary"),
          tooltip: {{ trigger: "axis", valueFormatter: value => Number(value).toFixed(0) + " lines" }},
          grid: {{ left: 64, right: 18, top: 30, bottom: 30 }},
          xAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => Math.round(value) }} }}),
          yAxis: axisStyle({{ type: "category", data: totals.map(row => row[0]) }}),
          series: [{{ type: "bar", data: totals.map(row => row[1]), itemStyle: {{ color: (params) => ordinalColor(totals, params.dataIndex) }}, barMaxWidth: 30 }}]
        }};
      }},
      fulfillmentPerf: function(type, filteredRows) {{
        const rows = filteredRows("daily");
        return {{
          ...chartBase("primary", "tertiary", "secondary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0 }},
          grid: {{ left: 52, right: 18, top: 30, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.date), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => Math.round(value) }} }}),
          series: [
            {{ name: "Labels", type: type, smooth: true, data: rows.map(row => row.fulfill_total), itemStyle: {{ color: chartTheme().muted }} }},
            {{ name: "Delivered", type: type, smooth: true, data: rows.map(row => row.fulfill_delivered), itemStyle: {{ color: chartTheme().secondary }}, areaStyle: type === "line" ? {{ opacity: 0.08 }} : undefined }},
            {{ name: "Exceptions", type: type, smooth: true, data: rows.map(row => row.fulfill_exceptions), itemStyle: {{ color: chartTheme().tertiary }} }}
          ]
        }};
      }},
      returnReasons: function(type, filteredRows) {{
        const totals = aggregateBy(filteredRows("daily"), "return_reasons");
        if (type === "pie") {{
          return {{
            ...chartBase(),
            color: totals.map((row, index) => categoricalColor(row[0], index)),
            tooltip: {{ trigger: "item", valueFormatter: value => Number(value).toFixed(0) + " returns" }},
            legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0, type: "scroll" }},
            series: [{{ type: "pie", radius: ["40%", "70%"], data: totals.map(row => ({{ name: row[0], value: row[1] }})) }}]
          }};
        }}
        return {{
          ...chartBase("tertiary"),
          tooltip: {{ trigger: "axis", valueFormatter: value => Number(value).toFixed(0) + " returns" }},
          grid: {{ left: 80, right: 18, top: 30, bottom: 30 }},
          xAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => Math.round(value) }} }}),
          yAxis: axisStyle({{ type: "category", data: totals.map(row => row[0]) }}),
          series: [{{ type: "bar", data: totals.map(row => row[1]), itemStyle: {{ color: (params) => categoricalColor(totals[params.dataIndex][0], params.dataIndex) }}, barMaxWidth: 28 }}]
        }};
      }},
      warehouseLoad: function(type, filteredRows) {{
        const totals = aggregateBy(filteredRows("daily"), "warehouses");
        // use warehouse dataset for richer profile, but ensure filter applies (re-aggregate from filtered daily for inventory)
        const inventory = {{}};
        const inbound = {{}};
        const unsellable = {{}};
        const fulfillment = {{}};
        const exceptions = {{}};
        filteredRows("daily").forEach(row => {{
          const warehouses = ((row && row.warehouses_active) || 0);
          // We only have aggregated inventory per day; surface a per-warehouse profile using most-recent day only
        }});
        // Fallback to non-filtered warehouse dataset; warehouse data is snapshot-level and does not change with range.
        const rows = (dashboardPayload.datasets.warehouse || []).map(row => ({{
          warehouse: row.warehouse,
          sellable: row.inventory,
          inbound: row.inbound,
          unsellable: row.unsellable,
          delivered: row.delivered,
          exceptions: row.exceptions
        }}));
        return {{
          ...chartBase("primary", "secondary", "tertiary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0 }},
          grid: {{ left: 70, right: 18, top: 30, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.warehouse), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => Math.round(value / 1000) + "k" }} }}),
          series: [
            {{ name: "Sellable", type: "bar", stack: "inv", data: rows.map(row => row.sellable), itemStyle: {{ color: chartTheme().primary }} }},
            {{ name: "Inbound", type: "bar", stack: "inv", data: rows.map(row => row.inbound), itemStyle: {{ color: chartTheme().secondary }} }},
            {{ name: "Unsellable", type: "line", data: rows.map(row => row.unsellable), smooth: true, itemStyle: {{ color: chartTheme().tertiary }} }},
            {{ name: "Exceptions", type: "line", data: rows.map(row => row.exceptions), smooth: true, itemStyle: {{ color: chartTheme().quaternary }} }}
          ]
        }};
      }},
      inventoryHealth: function(type, filteredRows) {{
        const rows = filteredRows("daily");
        return {{
          ...chartBase("primary", "secondary", "tertiary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0 }},
          grid: {{ left: 60, right: 60, top: 30, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.date), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: [
            axisStyle({{ type: "value", name: "Units", nameTextStyle: {{ color: chartTheme().muted }}, axisLabel: {{ formatter: value => Math.round(value / 1000) + "k" }} }}),
            axisStyle({{ type: "value", name: "Unsellable %", nameTextStyle: {{ color: chartTheme().muted }}, position: "right", splitLine: {{ show: false }}, axisLabel: {{ formatter: value => Math.round(value * 100) + "%" }}, max: 1 }}),
          ],
          series: [
            {{ name: "Sellable", type: type, smooth: true, data: rows.map(row => row.inventory_sellable), itemStyle: {{ color: chartTheme().primary }}, areaStyle: type === "line" ? {{ opacity: 0.08 }} : undefined }},
            {{ name: "Inbound", type: type, smooth: true, data: rows.map(row => row.inventory_inbound), itemStyle: {{ color: chartTheme().secondary }} }},
            {{
              name: "Unsellable %",
              type: type,
              smooth: true,
              yAxisIndex: 1,
              data: rows.map(row => {{
                const total = row.inventory_sellable + row.inventory_unsellable;
                return total ? row.inventory_unsellable / total : 0;
              }}),
              itemStyle: {{ color: chartTheme().tertiary }}
            }}
          ]
        }};
      }},
      dailyReturnCost: function(type, filteredRows) {{
        const rows = filteredRows("daily");
        return {{
          ...chartBase("tertiary", "primary", "quaternary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: chartTheme().text }}, top: 0 }},
          grid: {{ left: 60, right: 60, top: 30, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.date), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: [
            axisStyle({{ type: "value", name: "Returns", nameTextStyle: {{ color: chartTheme().muted }}, axisLabel: {{ formatter: value => Math.round(value) }} }}),
            axisStyle({{ type: "value", name: "Refund $", nameTextStyle: {{ color: chartTheme().muted }}, position: "right", splitLine: {{ show: false }}, axisLabel: {{ formatter: value => "$" + Math.round(value) }} }})
          ],
          series: [
            {{ name: "Returns", type: type, smooth: true, data: rows.map(row => row.returns), itemStyle: {{ color: chartTheme().tertiary }} }},
            {{ name: "Refund amount", type: type, smooth: true, yAxisIndex: 1, data: rows.map(row => row.refund_amount), itemStyle: {{ color: chartTheme().primary }} }},
            {{ name: "Failed refunds", type: type, smooth: true, data: rows.map(row => row.refund_failed), itemStyle: {{ color: chartTheme().quaternary }} }}
          ]
        }};
      }},
      anomalyMarkers: function(type, filteredRows) {{
        // Combine baseline magnitude + per-day revenue markers
        const rows = filteredRows("daily");
        const revenueSeries = rows.map(row => [row.date, row.revenue]);
        const anomalyIndex = new Set((dashboardPayload.anomalies || []).map(a => a.date));
        const anomalyMarkers = dashboardPayload.anomalies || [];
        if (type === "bar") {{
          return {{
            ...chartBase("tertiary"),
            tooltip: {{ trigger: "axis" }},
            grid: {{ left: 70, right: 30, top: 30, bottom: 36 }},
            xAxis: axisStyle({{ type: "category", data: anomalyMarkers.map(a => a.date), axisLabel: {{ hideOverlap: true }} }}),
            yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => Number(value).toFixed(1) }} }}),
            series: [{{ type: "bar", data: anomalyMarkers.map(a => a.magnitude), itemStyle: {{ color: chartTheme().tertiary }}, barMaxWidth: 18 }}]
          }};
        }}
        return {{
          ...chartBase("primary"),
          tooltip: {{
            trigger: "item",
            formatter: function (params) {{
              if (!params || !params.value) return "";
              if (Array.isArray(params.value)) {{
                return "{{date: " + params.value[0] + ", revenue: $" + Number(params.value[1]).toFixed(0) + "}}";
              }}
              return "Magnitude: " + Number(params.value[1]).toFixed(2);
            }}
          }},
          grid: {{ left: 60, right: 30, top: 30, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.date), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: value => "$" + Math.round(value / 1000) + "k" }} }}),
          series: [
            {{ type: "scatter", symbolSize: 14, data: anomalyMarkers.map(a => [a.date, a.revenue]), itemStyle: {{ color: chartTheme().tertiary }} }},
            {{ type: "line", smooth: true, data: rows.map(row => row.revenue), lineStyle: {{ color: chartTheme().primary, opacity: 0.5, width: 2 }}, symbol: "none", areaStyle: {{ color: chartTheme().primary, opacity: 0.08 }} }}
          ]
        }};
      }}
    }};
    const sourceMap = {json_script(source_map)};
    setupDashboardRuntime({{
      datasets: dashboardPayload.datasets,
      availableDates: dashboardPayload.availableDates,
      defaultRange: dashboardPayload.defaultRange,
      initialCharts: {json_script(initial_charts)},
      chartFactories,
      sourceMap,
      tables: {json_script(table_config)},
      fullScript: {js_string(analysis_logic)},
      modalSubtitlePrefix: "Operations panel transform for "
    }});
    """

    return f"""<!-- Generated by Trae Work -->
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(payload["title"])}</title>
  <style>{CSS}</style>
</head>
<body>
  <header class="topbar">
    <div class="topbar-inner">
      <div>
        <h1>{html.escape(payload["title"])}</h1>
        <p class="subtitle">{html.escape(payload["subtitle"])}</p>
        <p class="freshness" id="dataFreshness">Latest data: {html.escape(payload["freshness"]["latestDataDate"])} | Captured: {html.escape(payload["freshness"]["latestCapturedAt"])} | {html.escape(payload["timezone"])}</p>
      </div>
      <div class="controls" aria-label="Dashboard time controls">
        <span class="range-label" id="activeRangeLabel"></span>
        <div class="segmented" aria-label="Time preset">
          <button data-range-preset="7D">7D</button>
          <button data-range-preset="30D">30D</button>
          <button data-range-preset="MTD">MTD</button>
          <button data-range-preset="QTD">QTD</button>
          <button data-range-preset="ALL">All</button>
        </div>
        <div class="date-fields">
          <input id="rangeStart" data-range-input type="date" aria-label="Start date">
          <input id="rangeEnd" data-range-input type="date" aria-label="End date">
        </div>
        <div class="segmented theme-switch" aria-label="Theme">
          <button data-theme-choice="light" type="button" aria-label="Light theme" title="Light">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
              <circle cx="12" cy="12" r="4"></circle>
              <path d="M12 2v2"></path>
              <path d="M12 20v2"></path>
              <path d="m4.93 4.93 1.41 1.41"></path>
              <path d="m17.66 17.66 1.41 1.41"></path>
              <path d="M2 12h2"></path>
              <path d="M20 12h2"></path>
              <path d="m6.34 17.66-1.41 1.41"></path>
              <path d="m19.07 4.93-1.41 1.41"></path>
            </svg>
          </button>
          <button data-theme-choice="trae-dark" type="button" aria-label="Dark theme" title="Dark">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
              <path d="M20.99 13.53A8.5 8.5 0 1 1 10.47 3.01 7 7 0 0 0 20.99 13.53Z"></path>
            </svg>
          </button>
        </div>
      </div>
    </div>
  </header>
  <main class="dashboard-shell">
    {content}
  </main>
  <div id="modalBackdrop" class="modal-backdrop" role="dialog" aria-modal="true">
    <section class="modal">
      <div class="modal-head">
        <div>
          <h3 id="modalTitle">Data Source</h3>
          <p class="modal-subtitle" id="modalSubtitle"></p>
        </div>
        <button class="close" aria-label="Close" onclick="closeModal()">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
            <path d="M18 6 6 18"></path>
            <path d="m6 6 12 12"></path>
          </svg>
        </button>
      </div>
      <div class="modal-body">
        <section class="source-section">
          <h4>Panel transform</h4>
          <div class="code-wrap">
            <button class="copy-button" aria-label="Copy panel transform" onclick="copyCode('modalSnippet', this)">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
                <rect x="9" y="9" width="11" height="11" rx="2"></rect>
                <path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path>
              </svg>
            </button>
            <pre><code id="modalSnippet"></code></pre>
          </div>
        </section>
        <section class="source-section">
          <h4>Analysis logic</h4>
          <div class="code-wrap">
            <button class="copy-button" aria-label="Copy analysis logic" onclick="copyCode('modalCode', this)">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
                <rect x="9" y="9" width="11" height="11" rx="2"></rect>
                <path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path>
              </svg>
            </button>
            <pre><code id="modalCode"></code></pre>
          </div>
        </section>
      </div>
    </section>
  </div>
  <script>{echarts}</script>
  <script>{runtime}</script>
  <script>{chart_js}</script>
</body>
</html>
"""


def main() -> None:
    sheets = read_sources()
    payload = make_dashboard_payload(sheets)
    DASHBOARD_DATA.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    DASHBOARD_HTML.write_text(build_html(payload), encoding="utf-8")
    print(f"Wrote {DASHBOARD_HTML}")
    print(f"Wrote {DASHBOARD_DATA}")


if __name__ == "__main__":
    main()
