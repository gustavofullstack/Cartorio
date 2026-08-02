#!/usr/bin/env python3
"""User-Review Operating Dashboard.

Adapts dashboard_template.py to the uploaded `data/source_user_review.xlsx`,
which contains three sheets:

  - product_review_export    (520 reviews, ratings + verified + photo/video)
  - support_ticket_export    (360 tickets, source/category/CSAT)
  - marketplace_qa_export    (190 marketplace Q&A items, answer funnel)

The dashboard renders a single index.html with KPI tiles, ECharts panels,
a top-products leaderboard, and a freshness indicator. Time controls filter
all charts on the analytical date field (review_time, created_at, or
question_time depending on the panel).
"""

from __future__ import annotations

import csv
import html
import json
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
SNAPSHOT_DIR = ROOT / "data" / "snapshots"
SOURCE_XLSX = ROOT / "data" / "source_user_review.xlsx"
ECHARTS_JS = ROOT / "echarts.min.js"
DASHBOARD_RUNTIME_JS = ROOT / "dashboard_runtime.js"
DASHBOARD_HTML = ROOT / "index.html"
DASHBOARD_DATA = ROOT / "dashboard_data.json"

DASHBOARD_TITLE = "User Review & Support Operations"
DASHBOARD_SUBTITLE = (
    "Reviews · Support Tickets · Marketplace Q&A — refreshed from "
    "data/source_user_review.xlsx"
)
TIMEZONE_LABEL = "America/Sao_Paulo"
DEFAULT_RANGE = "30D"


# ---------- date helpers ----------

def excel_serial_to_date(serial: float | int | None) -> str | None:
    """Convert Excel serial date (days since 1899-12-30) to ISO YYYY-MM-DD.

    Daily precision only; fractional time-of-day is dropped on purpose so we
    never apply timezone shifts (the template contract forbids
    toISOString().slice(0,10) conversions).
    """
    if serial is None:
        return None
    try:
        days = int(float(serial))
    except (TypeError, ValueError):
        return None
    base = date(1899, 12, 30)
    return (base + timedelta(days=days)).isoformat()


def safe_int(value, default=0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def safe_float(value, default=0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


# ---------- snapshot I/O ----------

def read_xlsx_rows() -> list[dict]:
    """Read the three sheets of source_user_review.xlsx into a long shape.

    Each row represents one event with: `date`, `metric`, `segment`,
    `value`, plus per-event extras (rating, platform, sku, country, etc.).
    `metric` names the domain so the dashboard can pivot naturally.
    """
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    out: list[dict] = []

    # product reviews
    if "product_review_export" in wb.sheetnames:
        ws = wb["product_review_export"]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(headers, row))
            d = excel_serial_to_date(rec.get("review_time"))
            if not d:
                continue
            helpful = safe_int(rec.get("helpful_votes"))
            has_photo = bool(rec.get("has_photo"))
            has_video = bool(rec.get("has_video"))
            verified = bool(rec.get("verified_purchase"))
            rating = safe_int(rec.get("rating"))
            out.append(
                {
                    "date": d,
                    "metric": "review",
                    "segment": rec.get("source_platform") or "Unknown",
                    "value": 1,
                    "rating": rating,
                    "country": rec.get("country"),
                    "language": rec.get("language"),
                    "sku": rec.get("sku"),
                    "product_name": rec.get("product_name"),
                    "verified": verified,
                    "has_photo": has_photo,
                    "has_video": has_video,
                    "helpful_votes": helpful,
                    "moderation_status": rec.get("moderation_status"),
                    "seller_replied": rec.get("seller_reply_time") is not None,
                }
            )

    # support tickets
    if "support_ticket_export" in wb.sheetnames:
        ws = wb["support_ticket_export"]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(headers, row))
            d = excel_serial_to_date(rec.get("created_at"))
            if not d:
                continue
            csat = rec.get("csat_score")
            out.append(
                {
                    "date": d,
                    "metric": "ticket",
                    "segment": rec.get("source") or "Unknown",
                    "value": 1,
                    "ticket_category": rec.get("ticket_category"),
                    "ticket_status": rec.get("ticket_status"),
                    "refund_requested": bool(rec.get("refund_requested")),
                    "country": rec.get("country"),
                    "language": rec.get("language"),
                    "agent_id": rec.get("agent_id"),
                    "sku": rec.get("related_sku"),
                    "csat_score": safe_int(csat, default=0) if csat is not None else None,
                }
            )

    # marketplace Q&A
    if "marketplace_qa_export" in wb.sheetnames:
        ws = wb["marketplace_qa_export"]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = dict(zip(headers, row))
            d = excel_serial_to_date(rec.get("question_time"))
            if not d:
                continue
            answered = rec.get("answer_time") is not None
            out.append(
                {
                    "date": d,
                    "metric": "qa",
                    "segment": rec.get("platform") or "Unknown",
                    "value": 1,
                    "qa_answered": answered,
                    "answered_by": rec.get("answered_by"),
                    "visibility_status": rec.get("visibility_status"),
                    "upvotes": safe_int(rec.get("upvotes")),
                    "country": rec.get("country"),
                    "sku": rec.get("sku"),
                    "product_name": None,
                }
            )
    return out


def read_json_snapshot(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data
    rows = data.get("rows", [])
    return rows


def read_jsonl_snapshot(path: Path) -> list[dict]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def read_csv_snapshot(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def read_sources() -> list[dict]:
    rows: list[dict] = []
    if SOURCE_XLSX.exists():
        rows.extend(read_xlsx_rows())
    snapshots = sorted(SNAPSHOT_DIR.glob("*")) if SNAPSHOT_DIR.exists() else []
    for path in snapshots:
        if path.suffix.lower() == ".json":
            rows.extend(read_json_snapshot(path))
        elif path.suffix.lower() == ".jsonl":
            rows.extend(read_jsonl_snapshot(path))
        elif path.suffix.lower() == ".csv":
            rows.extend(read_csv_snapshot(path))
    return rows


def normalize_snapshots(rows: list[dict]) -> list[dict]:
    normalized: list[dict] = []
    for raw in rows:
        row = dict(raw)
        d = row.get("date") or excel_serial_to_date(row.get("raw_date"))
        d = str(d or "")[:10]
        if not d:
            continue
        row["date"] = d
        row["metric"] = str(row.get("metric") or "value")
        row["segment"] = str(row.get("segment") or "Unknown")
        row["value"] = safe_float(row.get("value"))
        normalized.append(row)
    return sorted(normalized, key=lambda r: (r["date"], r["metric"], r["segment"]))


def latest_date(rows: list[dict]) -> str:
    dates = sorted({r["date"] for r in rows if r.get("date")})
    return dates[-1] if dates else ""


def fmt_num(value: float) -> str:
    return f"{value:,.0f}"


def pct(value: float) -> str:
    return f"{value * 100:+.1f}%"


# ---------- aggregations ----------

def daily_metrics(rows: list[dict]) -> list[dict]:
    """Day-by-day counts of reviews, tickets, Q&A + derived NPS-ish score."""
    bucket: dict[str, dict] = defaultdict(
        lambda: {
            "date": "",
            "reviews": 0,
            "tickets": 0,
            "qa_questions": 0,
            "qa_answered": 0,
            "avg_rating": 0.0,
            "_rating_sum": 0.0,
            "_rating_n": 0,
            "csat_sum": 0.0,
            "_csat_n": 0,
            "helpful_votes": 0,
        }
    )
    for r in rows:
        slot = bucket[r["date"]]
        slot["date"] = r["date"]
        metric = r.get("metric")
        if metric == "review":
            slot["reviews"] += 1
            rating = r.get("rating")
            if rating is not None:
                slot["_rating_sum"] += rating
                slot["_rating_n"] += 1
            slot["helpful_votes"] += int(r.get("helpful_votes") or 0)
        elif metric == "ticket":
            slot["tickets"] += 1
            csat = r.get("csat_score")
            if csat:
                slot["csat_sum"] += csat
                slot["_csat_n"] += 1
        elif metric == "qa":
            slot["qa_questions"] += 1
            if r.get("qa_answered"):
                slot["qa_answered"] += 1
    out = []
    for key in sorted(bucket):
        s = bucket[key]
        s["avg_rating"] = (s["_rating_sum"] / s["_rating_n"]) if s["_rating_n"] else 0
        s["csat_avg"] = (s["csat_sum"] / s["_csat_n"]) if s["_csat_n"] else 0
        s.pop("_rating_sum", None)
        s.pop("_rating_n", None)
        s.pop("csat_sum", None)
        s.pop("_csat_n", None)
        out.append(s)
    return out


def rating_distribution(rows: list[dict]) -> list[dict]:
    counts = Counter()
    for r in rows:
        if r.get("metric") == "review" and r.get("rating") is not None:
            counts[int(r["rating"])] += 1
    return [{"rating": k, "count": counts.get(k, 0)} for k in [1, 2, 3, 4, 5]]


def platform_volume(rows: list[dict]) -> list[dict]:
    by_seg: dict[str, dict] = defaultdict(
        lambda: {"segment": "", "reviews": 0, "tickets": 0, "qa": 0}
    )
    for r in rows:
        seg = r.get("segment") or "Unknown"
        slot = by_seg[seg]
        slot["segment"] = seg
        m = r.get("metric")
        if m == "review":
            slot["reviews"] += 1
        elif m == "ticket":
            slot["tickets"] += 1
        elif m == "qa":
            slot["qa"] += 1
    out = list(by_seg.values())
    for slot in out:
        slot["total"] = slot["reviews"] + slot["tickets"] + slot["qa"]
    out.sort(key=lambda x: x["total"], reverse=True)
    return out


def ticket_category_mix(rows: list[dict]) -> list[dict]:
    counts: Counter = Counter()
    for r in rows:
        if r.get("metric") == "ticket":
            counts[r.get("ticket_category") or "Unknown"] += 1
    total = sum(counts.values()) or 1
    out = [{"category": k, "count": v, "share": v / total} for k, v in counts.items()]
    out.sort(key=lambda x: x["count"], reverse=True)
    return out


def qa_funnel(rows: list[dict]) -> list[dict]:
    total = 0
    answered = 0
    visible = 0
    pending = 0
    flagged = 0
    hidden = 0
    for r in rows:
        if r.get("metric") != "qa":
            continue
        total += 1
        if r.get("qa_answered"):
            answered += 1
        vs = r.get("visibility_status")
        if vs == "visible":
            visible += 1
        elif vs == "pending":
            pending += 1
        elif vs == "flagged":
            flagged += 1
        elif vs == "hidden":
            hidden += 1
    return [
        {"stage": "Asked", "count": total},
        {"stage": "Answered", "count": answered},
        {"stage": "Visible", "count": visible},
        {"stage": "Pending", "count": pending},
        {"stage": "Flagged", "count": flagged},
        {"stage": "Hidden", "count": hidden},
    ]


def product_leaderboard(rows: list[dict]) -> list[dict]:
    by_sku: dict[str, dict] = defaultdict(
        lambda: {
            "sku": "",
            "product_name": "",
            "reviews": 0,
            "tickets": 0,
            "rating_sum": 0.0,
            "rating_n": 0,
            "helpful": 0,
        }
    )
    for r in rows:
        sku = r.get("sku") or "Unknown"
        slot = by_sku[sku]
        slot["sku"] = sku
        slot["product_name"] = r.get("product_name") or slot["product_name"] or sku
        if r.get("metric") == "review":
            slot["reviews"] += 1
            rating = r.get("rating")
            if rating:
                slot["rating_sum"] += rating
                slot["rating_n"] += 1
            slot["helpful"] += int(r.get("helpful_votes") or 0)
        elif r.get("metric") == "ticket":
            slot["tickets"] += 1
    out = []
    for slot in by_sku.values():
        avg = (slot["rating_sum"] / slot["rating_n"]) if slot["rating_n"] else 0
        out.append(
            {
                "sku": slot["sku"],
                "product_name": slot["product_name"],
                "reviews": slot["reviews"],
                "tickets": slot["tickets"],
                "avg_rating": round(avg, 2),
                "helpful": slot["helpful"],
            }
        )
    out.sort(key=lambda x: x["reviews"], reverse=True)
    return out


def geography_table(rows: list[dict]) -> list[dict]:
    bucket: dict[str, dict] = defaultdict(
        lambda: {
            "country": "",
            "reviews": 0,
            "tickets": 0,
            "qa": 0,
            "rating_sum": 0.0,
            "rating_n": 0,
        }
    )
    for r in rows:
        country = r.get("country") or "Unknown"
        slot = bucket[country]
        slot["country"] = country
        m = r.get("metric")
        if m == "review":
            slot["reviews"] += 1
            rating = r.get("rating")
            if rating:
                slot["rating_sum"] += rating
                slot["rating_n"] += 1
        elif m == "ticket":
            slot["tickets"] += 1
        elif m == "qa":
            slot["qa"] += 1
    out = []
    for slot in bucket.values():
        avg = (slot["rating_sum"] / slot["rating_n"]) if slot["rating_n"] else 0
        out.append(
            {
                "country": slot["country"],
                "reviews": slot["reviews"],
                "tickets": slot["tickets"],
                "qa": slot["qa"],
                "avg_rating": round(avg, 2),
            }
        )
    out.sort(key=lambda x: x["reviews"] + x["tickets"], reverse=True)
    return out


# ---------- payload ----------

def make_dashboard_payload(rows: list[dict]) -> dict:
    daily = daily_metrics(rows)
    dates = [item["date"] for item in daily]
    latest = dates[-1] if dates else ""
    first = dates[0] if dates else ""

    def window(days: int) -> tuple[str, str]:
        if not dates:
            return ("", "")
        end_idx = len(dates) - 1
        start_idx = max(0, end_idx - days + 1)
        return (dates[start_idx], dates[end_idx])

    cur_start, cur_end = window(min(30, len(dates)))
    prev_start, prev_end = window(min(60, len(dates)))

    def totals(start: str, end: str) -> dict:
        n_reviews = n_tickets = n_qa = ans = rate_sum = rate_n = csat_sum = csat_n = 0
        for r in rows:
            d = r["date"]
            if start and (d < start or d > end):
                continue
            if r.get("metric") == "review":
                n_reviews += 1
                rating = r.get("rating")
                if rating:
                    rate_sum += rating
                    rate_n += 1
            elif r.get("metric") == "ticket":
                n_tickets += 1
                cs = r.get("csat_score")
                if cs:
                    csat_sum += cs
                    csat_n += 1
            elif r.get("metric") == "qa":
                n_qa += 1
                if r.get("qa_answered"):
                    ans += 1
        return {
            "reviews": n_reviews,
            "tickets": n_tickets,
            "qa": n_qa,
            "answered": ans,
            "avg_rating": (rate_sum / rate_n) if rate_n else 0,
            "csat_avg": (csat_sum / csat_n) if csat_n else 0,
        }

    cur = totals(cur_start, cur_end)
    # Use the row count of `daily` to detect whether a prior period exists.
    if prev_start == cur_start and prev_end == cur_end:
        prev = {"reviews": 0, "tickets": 0, "qa": 0, "answered": 0,
                "avg_rating": 0, "csat_avg": 0}
    else:
        prev = totals(prev_start, prev_end)

    def delta(curr: float, prior: float) -> float:
        return ((curr - prior) / prior) if prior else 0.0

    tickets_share = (
        (cur["answered"] / cur["qa"]) if cur["qa"] else 0.0
    )  # safe re-use: this is qa answer-rate; overwritten below
    # proper qa answer rate
    qa_answer_rate = (cur["answered"] / cur["qa"]) if cur["qa"] else 0
    prev_qa_answer_rate = (prev["answered"] / prev["qa"]) if prev["qa"] else 0

    latest_captured = max(
        (row.get("captured_at", "") for row in rows), default=""
    )
    if not latest_captured:
        latest_captured = datetime.now(timezone.utc).isoformat(timespec="seconds")

    source_snippets = {
        "reviewTrend": (
            "# daily review volume + avg rating derived from product_review_export\n"
            "daily = daily_metrics(rows)\n"
            "filtered = [r for r in daily if start_date <= r['date'] <= end_date]\n"
            "x = [r['date'] for r in filtered]\n"
            "reviews = [r['reviews'] for r in filtered]\n"
            "avg_rating = [round(r['avg_rating'], 2) for r in filtered]"
        ),
        "ratingMix": (
            "# rating breakdown over the active window\n"
            "review_rows = [r for r in rows if r['metric'] == 'review']\n"
            "filtered = [r for r in review_rows if start_date <= r['date'] <= end_date]\n"
            "counts = Counter(int(r['rating']) for r in filtered if r.get('rating'))\n"
            "return [{'rating': k, 'count': counts.get(k, 0)} for k in [1,2,3,4,5]]"
        ),
        "platformVolume": (
            "# volume by source platform (reviews + tickets + qa combined)\n"
            "rows = rows  # all three sheets merged\n"
            "filtered = [r for r in rows if start_date <= r['date'] <= end_date]\n"
            "groups = Counter(r['segment'] for r in filtered)\n"
            "stacked = pivot on metric (review / ticket / qa) per segment"
        ),
        "ticketCategories": (
            "# support ticket category breakdown from support_ticket_export\n"
            "ticket_rows = [r for r in rows if r['metric'] == 'ticket'\n"
            "                and start_date <= r['date'] <= end_date]\n"
            "counter = Counter(r['ticket_category'] for r in ticket_rows)\n"
            "return [{'category': k, 'count': v} for k, v in counter.items()]"
        ),
        "qaFunnel": (
            "# marketplace Q&A funnel from marketplace_qa_export\n"
            "qa_rows = [r for r in rows if r['metric'] == 'qa'\n"
            "            and start_date <= r['date'] <= end_date]\n"
            "asked = len(qa_rows)\n"
            "answered = sum(1 for r in qa_rows if r['qa_answered'])\n"
            "visibility = Counter(r['visibility_status'] for r in qa_rows)"
        ),
        "topProducts": (
            "# leaderboard of SKUs by review volume + ticket load\n"
            "by_sku = defaultdict()\n"
            "for r in rows:\n"
            "    sku = r['sku']\n"
            "    if r['metric'] == 'review': by_sku[sku]['reviews'] += 1\n"
            "    if r['metric'] == 'ticket': by_sku[sku]['tickets'] += 1\n"
            "return sorted(rows, key=lambda r: r['reviews'], reverse=True)"
        ),
        "geography": (
            "# country rollup across reviews / tickets / qa\n"
            "rows_by_country = group rows by r['country'] after time filter\n"
            "metrics per country: review count, ticket count, qa count, avg rating"
        ),
    }

    return {
        "title": DASHBOARD_TITLE,
        "subtitle": DASHBOARD_SUBTITLE,
        "timezone": TIMEZONE_LABEL,
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "freshness": {
            "latestDataDate": latest,
            "earliestDataDate": first,
            "latestCapturedAt": latest_captured,
            "source": "data/source_user_review.xlsx (3 sheets, immutable)",
            "snapshotDate": latest,
        },
        "availableDates": dates,
        "defaultRange": DEFAULT_RANGE,
        "kpis": [
            {
                "id": "reviews30",
                "label": "Reviews (window)",
                "value": fmt_num(cur["reviews"]),
                "delta": pct(delta(cur["reviews"], prev["reviews"])),
                "detail": f"vs prior {len(dates)//2 or len(dates)}-day window" if prev["reviews"] else "no prior window",
            },
            {
                "id": "rating30",
                "label": "Avg review rating",
                "value": f"{cur['avg_rating']:.2f} / 5",
                "delta": (
                    f"{(cur['avg_rating'] - prev['avg_rating']):+.2f}"
                    if prev["avg_rating"]
                    else "n/a"
                ),
                "detail": "1–5 star, review-only",
            },
            {
                "id": "tickets30",
                "label": "Support tickets",
                "value": fmt_num(cur["tickets"]),
                "delta": pct(delta(cur["tickets"], prev["tickets"])),
                "detail": (
                    f"CSAT {cur['csat_avg']:.2f}/5"
                    if cur["csat_avg"]
                    else "CSAT n/a"
                ),
            },
            {
                "id": "qa30",
                "label": "Marketplace Q&A",
                "value": fmt_num(cur["qa"]),
                "delta": f"{qa_answer_rate * 100:.1f}% answered",
                "detail": f"{cur['answered']} answered / {cur['qa']} asked",
            },
        ],
        "datasets": {
            "daily": daily,
            "ratings": rating_distribution(rows),
            "platforms": platform_volume(rows),
            "ticketCategories": ticket_category_mix(rows),
            "qaFunnel": qa_funnel(rows),
            "topProducts": product_leaderboard(rows),
            "geography": geography_table(rows),
        },
        "sourceSnippets": source_snippets,
    }


# ---------- formatting helpers ----------

def js_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False).replace("</", "<\\/")


def json_script(value) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace(
        "</", "<\\/"
    )


def render_kpi_block(block: dict) -> str:
    return f"""
    <section class="kpi-tile" id="{html.escape(block['id'])}">
      <p>{html.escape(block['label'])}</p>
      <strong>{html.escape(block['value'])}</strong>
      <span>{html.escape(block['delta'])}</span>
      <small>{html.escape(block['detail'])}</small>
    </section>
    """


def render_panel_actions(block: dict) -> str:
    edit_command = ""
    edit = ""
    if len(block.get("allowed_types", [])) > 1:
        options = "\n".join(
            f'<option value="{html.escape(kind)}"{" selected" if kind == block.get("initial_type") else ""}>{html.escape(kind)}</option>'
            for kind in block["allowed_types"]
        )
        edit_command = (
            f"<button onclick=\"toggleEdit('{html.escape(block['chart_id'])}')\">Edit</button>"
        )
        edit = f"""
        <div class="edit-panel" id="edit-{html.escape(block['chart_id'])}">
          <label for="select-{html.escape(block['chart_id'])}">Type</label>
          <select id="select-{html.escape(block['chart_id'])}" onchange="setChartType('{html.escape(block['chart_id'])}', this.value)">
            {options}
          </select>
        </div>
        """
    return f"""
    <div class="chart-actions">
      {edit}
      <div class="toolbox">
        <button class="tool-button" aria-label="Panel actions" onclick="toggleMenu('{html.escape(block['chart_id'])}')"><span class="dot"></span><span class="dot"></span><span class="dot"></span></button>
        <div class="menu" id="menu-{html.escape(block['chart_id'])}">
          {edit_command}
          <button onclick="viewSource('{html.escape(block['source_key'])}')">View Data Source</button>
        </div>
      </div>
    </div>
    """


def infer_panel_span(block: dict) -> int:
    if block.get("span") is not None:
        s = int(block["span"])
        return s if s in (4, 6, 12) else 6
    if block["kind"] == "table":
        cols = block.get("columns", [])
        has_long = any(c.get("long_text") for c in cols)
        return 12 if len(cols) >= 6 or has_long else 6
    if block["kind"] == "chart":
        chart_type = str(block.get("initial_type") or "")
        dense = chart_type in {"heatmap", "scatter", "funnel"}
        many = int(block.get("category_count") or 0) > 8
        return 12 if dense or many else 6
    if block["kind"] == "note":
        return 4 if block.get("compact") else 6
    return 6


def panel_span_attr(block: dict) -> str:
    return f'data-span="{infer_panel_span(block)}"'


def render_chart_block(block: dict) -> str:
    return f"""
    <section class="dashboard-panel chart-panel" {panel_span_attr(block)} id="{html.escape(block['id'])}">
      <header>
        <div>
          <h2>{html.escape(block['title'])}</h2>
          <p>{html.escape(block['subtitle'])}</p>
        </div>
        {render_panel_actions(block)}
      </header>
      <div class="chart" id="{html.escape(block['chart_id'])}" role="img" aria-label="{html.escape(block['title'])}"></div>
      <footer>{html.escape(block['unit'])} | {html.escape(block['source_context'])}</footer>
    </section>
    """


def render_table_block(block: dict) -> str:
    head = "".join(f"<th>{html.escape(col['label'])}</th>" for col in block["columns"])
    return f"""
    <section class="dashboard-panel table-panel" {panel_span_attr(block)} id="{html.escape(block['id'])}">
      <header>
        <div>
          <h2>{html.escape(block['title'])}</h2>
          <p>{html.escape(block['subtitle'])}</p>
        </div>
        <div class="toolbox">
          <button class="tool-button" aria-label="Panel actions" onclick="toggleMenu('{html.escape(block['source_key'])}')"><span class="dot"></span><span class="dot"></span><span class="dot"></span></button>
          <div class="menu" id="menu-{html.escape(block['source_key'])}">
            <button onclick="viewSource('{html.escape(block['source_key'])}')">View Data Source</button>
          </div>
        </div>
      </header>
      <div class="table-scroll">
        <table id="{html.escape(block['table_id'])}">
          <thead><tr>{head}</tr></thead>
          <tbody></tbody>
        </table>
      </div>
      <footer>{html.escape(block['source_context'])}</footer>
    </section>
    """


def render_note_block(block: dict) -> str:
    return f"""
    <section class="dashboard-note" {panel_span_attr(block)} id="{html.escape(block['id'])}">
      <strong>{html.escape(block['title'])}</strong>
      <span>{html.escape(block['body'])}</span>
    </section>
    """


# ---------- block composition ----------

def build_dashboard_blocks(payload: dict) -> list[dict]:
    return [
        # KPI tiles
        *[{"kind": "kpi", **kpi} for kpi in payload["kpis"]],
        # Charts
        {
            "kind": "chart",
            "id": "panel-review-trend",
            "chart_id": "reviewTrend",
            "source_key": "reviewTrend",
            "title": "Daily review volume · avg rating",
            "subtitle": "Reviews from product_review_export, smoothed weekly view",
            "unit": "reviews / day",
            "source_context": (
                f"Source: data/source_user_review.xlsx → product_review_export "
                f"({len(payload['availableDates'])} days)"
            ),
            "allowed_types": ["line", "bar"],
            "initial_type": "line",
        },
        {
            "kind": "chart",
            "id": "panel-rating-mix",
            "chart_id": "ratingMix",
            "source_key": "ratingMix",
            "title": "Rating distribution",
            "subtitle": "1–5 star share in active window",
            "unit": "reviews",
            "source_context": "Source: product_review_export, count by rating",
            "allowed_types": ["bar", "pie"],
            "initial_type": "bar",
        },
        {
            "kind": "chart",
            "id": "panel-platform-volume",
            "chart_id": "platformVolume",
            "source_key": "platformVolume",
            "title": "Volume by source platform",
            "subtitle": "Reviews + tickets + Q&A across Amazon, TikTok Shop, Shopify, Trustpilot, Walmart, Zendesk, Instagram, Buyer-Seller, Inbox",
            "unit": "events",
            "source_context": "Source: all 3 sheets, grouped by source_platform / source / platform",
            "allowed_types": ["bar", "line"],
            "initial_type": "bar",
            "category_count": 12,
        },
        {
            "kind": "chart",
            "id": "panel-ticket-categories",
            "chart_id": "ticketCategories",
            "source_key": "ticketCategories",
            "title": "Support ticket categories",
            "subtitle": "Where pressure is coming from (subscription / shipping / refund / damaged_item / return / payment / product_question)",
            "unit": "tickets",
            "source_context": "Source: support_ticket_export, ticket_category field",
            "allowed_types": ["bar", "pie"],
            "initial_type": "bar",
        },
        {
            "kind": "chart",
            "id": "panel-qa-funnel",
            "chart_id": "qaFunnel",
            "source_key": "qaFunnel",
            "title": "Marketplace Q&A funnel",
            "subtitle": "Asked → Answered → Visibility stages",
            "unit": "questions",
            "source_context": "Source: marketplace_qa_export, answer_time + visibility_status",
            "allowed_types": ["bar", "line"],
            "initial_type": "bar",
        },
        # Tables
        {
            "kind": "table",
            "id": "panel-top-products",
            "table_id": "topProductsTable",
            "source_key": "topProducts",
            "title": "Top products",
            "subtitle": "Sorted by review volume across all marketplaces",
            "source_context": "Source: product_review_export + support_ticket_export joined on SKU",
            "columns": [
                {"field": "product_name", "label": "Product"},
                {"field": "sku", "label": "SKU"},
                {"field": "reviews", "label": "Reviews", "numeric": True},
                {"field": "tickets", "label": "Tickets", "numeric": True},
                {"field": "avg_rating", "label": "Avg ★", "numeric": True},
                {"field": "helpful", "label": "Helpful votes", "numeric": True},
            ],
        },
        {
            "kind": "table",
            "id": "panel-geography",
            "table_id": "geographyTable",
            "source_key": "geography",
            "title": "Geography mix",
            "subtitle": "Country rollup across all event sources",
            "source_context": "Source: all 3 sheets, grouped by country",
            "columns": [
                {"field": "country", "label": "Country"},
                {"field": "reviews", "label": "Reviews", "numeric": True},
                {"field": "tickets", "label": "Tickets", "numeric": True},
                {"field": "qa", "label": "Q&A", "numeric": True},
                {"field": "avg_rating", "label": "Avg ★", "numeric": True},
            ],
        },
        # Note
        {
            "kind": "note",
            "id": "automation-note",
            "title": "Refresh & data lineage",
            "body": (
                "Source: data/source_user_review.xlsx (3 sheets). Re-run "
                "python dashboard.py to regenerate after the Excel is refreshed "
                "or after new JSON/JSONL/CSV snapshots land under "
                "data/snapshots/. KPI / chart time windows are derived locally "
                "from row.get('date') in YYYY-MM-DD; no UTC conversion."
            ),
        },
    ]


def render_dashboard_blocks(blocks: list[dict]) -> str:
    kpis = "\n".join(render_kpi_block(b) for b in blocks if b["kind"] == "kpi")
    panels = []
    for b in blocks:
        if b["kind"] == "chart":
            panels.append(render_chart_block(b))
        elif b["kind"] == "table":
            panels.append(render_table_block(b))
        elif b["kind"] == "note":
            panels.append(render_note_block(b))
    return f"""
    <section class="kpi-grid">{kpis}</section>
    <section class="panel-grid">{''.join(panels)}</section>
    """


ANALYSIS_LOGIC = """Analysis logic
- Source: data/source_user_review.xlsx — three immutable sheets:
    * product_review_export  (520 reviews)
    * support_ticket_export  (360 tickets)
    * marketplace_qa_export  (190 Q&A items)
- Excel serial dates (1899-12-30 base) are converted to local YYYY-MM-DD via
  excel_serial_to_date(). No toISOString().slice(0,10) conversion; timezone
  stays consistent across the payload.
- normalize_snapshots() trims to (date, metric, segment) grain and keeps
  `value=1` per event so simple counts are preserved while richer fields
  (rating, helpful_votes, csat_score, qa_answered, etc.) survive in extras.
- Aggregations:
    daily_metrics()      daily review / ticket / qa counts + avg rating / CSAT
    rating_distribution() bucket reviews by 1-5 stars
    platform_volume()    segment == source platform name (joined sheet)
    ticket_category_mix() Counter over support_ticket_export.ticket_category
    qa_funnel()          asked → answered → visibility pipeline
    product_leaderboard() per-SKU rollup (reviews + tickets + avg rating)
    geography_table()    per-country rollup across all events
- Time filtering: dashboard_runtime.js filters all rows by row.date against
  the active 7D/30D/MTD/QTD/YTD/ALL window. Switching the preset re-renders
  every chart and table; KPI delta vs prior window recomputes client-side
  using filtered row counts.
- Snapshots: append an immutable JSON/JSONL/CSV file to data/snapshots/ to
  layer new event history; rerun python dashboard.py to refresh the payload.
"""


def build_html(payload: dict) -> str:
    echarts = ECHARTS_JS.read_text(encoding="utf-8")
    runtime = DASHBOARD_RUNTIME_JS.read_text(encoding="utf-8")
    blocks = build_dashboard_blocks(payload)
    content = render_dashboard_blocks(blocks)
    initial_charts = [
        {"id": b["chart_id"], "type": b["initial_type"]}
        for b in blocks
        if b["kind"] == "chart"
    ]
    table_config = {
        "topProductsTable": {
            "dataset": "topProducts",
            "sortField": "reviews",
            "sortDirection": "desc",
            "limit": 12,
            "columns": [
                {"field": "product_name"},
                {"field": "sku"},
                {"field": "reviews", "numeric": True},
                {"field": "tickets", "numeric": True},
                {"field": "avg_rating", "numeric": True},
                {"field": "helpful", "numeric": True},
            ],
        },
        "geographyTable": {
            "dataset": "geography",
            "sortField": "reviews",
            "sortDirection": "desc",
            "limit": 12,
            "columns": [
                {"field": "country"},
                {"field": "reviews", "numeric": True},
                {"field": "tickets", "numeric": True},
                {"field": "qa", "numeric": True},
                {"field": "avg_rating", "numeric": True},
            ],
        },
    }
    source_map = payload["sourceSnippets"]

    css = """
    :root {
      color-scheme: light;
      --ink: #2f3437;
      --muted: #68707a;
      --faint: #8b95a3;
      --line: #e1e5ea;
      --line-strong: #cbd3dc;
      --panel: #FAFAFA;
      --page: #ffffff;
      --surface: #FAFAFA;
      --soft: #f2f3f5;
      --soft-blue: #f3f6fb;
      --control-bg: rgba(255, 255, 255, 0.94);
      --topbar-bg: rgba(255, 255, 255, 0.96);
      --menu-bg: #ffffff;
      --modal-bg: #ffffff;
      --modal-backdrop: rgba(55, 53, 47, 0.34);
      --table-head: #FAFAFA;
      --table-hover: #f1f2ff;
      --chart-bg: #FAFAFA;
      --chart-text: #2f3437;
      --chart-muted: #68707a;
      --chart-line: #e1e5ea;
      --chart-primary: #2F6BFF;
      --chart-secondary: #00BFA6;
      --chart-tertiary: #FF7A3D;
      --chart-quaternary: #F45BB3;
      --chart-1: #F45BB3;
      --chart-2: #2F6BFF;
      --chart-3: #00BFA6;
      --chart-4: #FF7A3D;
      --chart-5: #9BD82E;
      --chart-6: #7C3AED;
      --chart-7: #FFD23F;
      --brand: #6979F8;
      --brand-hover: #9EA9FF;
      --brand-end: #CDD2FD;
      --brand-text: #ffffff;
      --accent: #2F6BFF;
      --accent-2: #00BFA6;
      --warn: #b7791f;
    }
    html[data-theme="trae-dark"] {
      color-scheme: dark;
      --ink: #f5f9fe;
      --muted: #9599a6;
      --faint: #666b75;
      --line: #2a2d31;
      --line-strong: #3a3f45;
      --panel: #1a1b1d;
      --page: #0c0c0d;
      --surface: #222427;
      --soft: #2a2d31;
      --soft-blue: #202123;
      --control-bg: #202123;
      --topbar-bg: rgba(12, 12, 13, 0.92);
      --menu-bg: #202123;
      --modal-bg: #1a1b1d;
      --modal-backdrop: rgba(0, 0, 0, 0.58);
      --table-head: #222427;
      --table-hover: #202123;
      --chart-bg: #222427;
      --chart-text: #d1d3db;
      --chart-muted: #9599a6;
      --chart-line: #2a2d31;
      --chart-primary: #28d9ff;
      --chart-secondary: #32f08c;
      --chart-tertiary: #f6c85f;
      --chart-quaternary: #ff6b9a;
      --chart-1: #32f08c;
      --chart-2: #28d9ff;
      --chart-3: #a78bfa;
      --chart-4: #f6c85f;
      --chart-5: #ff6b9a;
      --chart-6: #6ea8ff;
      --chart-7: #d1d3db;
      --brand: #32f08c;
      --brand-hover: #0fdc78;
      --brand-end: #32f08c;
      --brand-text: #0c0c0d;
      --accent: #32f08c;
      --accent-2: #0fdc78;
    }
    * { box-sizing: border-box; }
    html { background: var(--page); }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Helvetica Neue", Arial, sans-serif;
      background: var(--page);
      color: var(--ink);
      font-size: 1rem;
      line-height: 1.55;
      -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility;
    }
    .topbar {
      position: sticky;
      top: 0;
      z-index: 20;
      border-bottom: 1px solid var(--line);
      background: var(--topbar-bg);
      backdrop-filter: blur(12px);
    }
    .topbar-inner {
      max-width: 1320px;
      margin: 0 auto;
      padding: 14px 22px;
      display: grid;
      grid-template-columns: minmax(260px, 1fr) auto;
      gap: 18px;
      align-items: center;
    }
    h1, h2, p { margin: 0; }
    h1 { font-size: 22px; font-weight: 500; letter-spacing: 0; }
    .subtitle, .freshness, .range-label, .dashboard-panel p, footer, small {
      color: var(--muted);
      font-size: 13px;
      line-height: 1.45;
      font-weight: 400;
    }
    .controls {
      display: flex;
      flex-wrap: wrap;
      align-items: center;
      justify-content: flex-end;
      gap: 10px;
    }
    .range-label { display: none; }
    .segmented {
      display: inline-flex;
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
      background: var(--control-bg);
    }
    .segmented button, .menu button, .edit-panel button {
      border: 0;
      background: transparent;
      color: var(--ink);
      font: inherit;
      cursor: pointer;
    }
    .segmented button {
      min-width: 44px;
      height: 34px;
      padding: 0 10px;
      border-right: 1px solid var(--line);
      font-size: 13px;
      font-weight: 400;
    }
    .segmented button:last-child { border-right: 0; }
    .segmented button.active { background: var(--brand); color: var(--brand-text); font-weight: 500; }
    .theme-switch button.active { background: var(--brand); color: var(--brand-text); }
    .theme-switch button {
      width: 38px;
      min-width: 38px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      padding: 0;
    }
    .theme-switch svg { width: 16px; height: 16px; stroke-width: 2; }
    .date-fields { display: inline-flex; align-items: center; gap: 6px; }
    input[type="date"] {
      height: 34px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 0 8px;
      background: var(--control-bg);
      color: var(--ink);
      font: inherit;
      font-size: 13px;
      font-weight: 400;
    }
    .dashboard-shell { max-width: 1320px; margin: 0 auto; padding: 18px 22px 44px; }
    .kpi-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }
    .kpi-tile {
      min-height: 126px;
      padding: 15px;
      display: grid;
      align-content: space-between;
      gap: 8px;
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 12px;
    }
    .kpi-tile:first-child {
      background: linear-gradient(135deg, var(--brand) 0%, var(--brand-hover) 58%, var(--brand-end) 100%);
      border-color: var(--brand);
    }
    .kpi-tile p { color: var(--muted); font-size: 13px; font-weight: 500; }
    .kpi-tile strong { font-size: 28px; font-weight: 500; letter-spacing: 0; }
    .kpi-tile span { color: var(--ink); font-size: 15px; font-weight: 500; line-height: 1.35; }
    .kpi-tile small { font-size: 13px; font-weight: 400; }
    .kpi-tile:first-child p,
    .kpi-tile:first-child strong,
    .kpi-tile:first-child span,
    .kpi-tile:first-child small { color: var(--brand-text); }
    .panel-grid {
      display: grid;
      grid-template-columns: repeat(12, minmax(0, 1fr));
      gap: 20px 16px;
    }
    .dashboard-panel {
      min-height: 360px;
      display: flex;
      flex-direction: column;
      gap: 10px;
      background: transparent;
      border: 0;
      border-radius: 0;
      padding: 0;
    }
    [data-span="4"] { grid-column: span 4; }
    [data-span="6"] { grid-column: span 6; }
    [data-span="12"] { grid-column: 1 / -1; }
    .dashboard-note {
      min-height: 180px;
      padding: 16px;
      display: flex;
      flex-direction: column;
      gap: 8px;
      border: 1px solid var(--line);
      border-radius: 16px;
      background: transparent;
    }
    .dashboard-panel header {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      min-height: 42px;
    }
    .dashboard-panel h2 { font-size: 17px; font-weight: 500; letter-spacing: 0; }
    .chart {
      width: 100%;
      height: 276px;
      min-height: 276px;
      padding: 8px 0 6px;
      border: 1px solid var(--line);
      border-radius: 16px;
      background: var(--chart-bg);
    }
    .chart-actions {
      display: flex;
      align-items: center;
      justify-content: flex-end;
      gap: 8px;
      flex: 0 0 auto;
      position: relative;
      z-index: 12;
    }
    .toolbox { position: relative; flex: 0 0 auto; }
    .tool-button {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 3px;
      width: 34px;
      height: 30px;
      border: 1px solid var(--line-strong);
      border-radius: 999px;
      background: var(--control-bg);
      color: var(--muted);
      cursor: pointer;
      font-size: 0;
      line-height: 0;
      padding: 0;
      opacity: 0;
      transition: opacity 140ms ease, background-color 140ms ease;
    }
    .tool-button .dot { display: block; width: 3px; height: 3px; border-radius: 50%; background: currentColor; }
    .dashboard-panel:hover .tool-button,
    .dashboard-panel:focus-within .tool-button,
    .dashboard-note:hover .tool-button,
    .dashboard-note:focus-within .tool-button { opacity: 1; }
    .menu {
      display: none;
      position: absolute;
      right: 0;
      top: 34px;
      z-index: 40;
      width: 188px;
      padding: 6px;
      border: 1px solid var(--line-strong);
      border-radius: 8px;
      background: var(--menu-bg);
      box-shadow: 0 8px 18px rgba(32, 33, 36, 0.12);
    }
    .menu.open { display: block; }
    .edit-panel {
      display: none;
      align-items: center;
      gap: 6px;
      height: 28px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--control-bg);
      color: var(--muted);
      font-size: 13px;
    }
    .edit-panel.open { display: flex; }
    .edit-panel label { padding-left: 8px; white-space: nowrap; }
    .edit-panel select {
      height: 26px;
      border: 0;
      border-left: 1px solid var(--line);
      border-radius: 0 5px 5px 0;
      padding: 0 24px 0 7px;
      background: transparent;
      color: var(--ink);
      font: inherit;
      font-size: 13px;
      outline: none;
      cursor: pointer;
    }
    .edit-panel select:focus-visible {
      outline: 2px solid var(--brand);
    }
    .menu button {
      display: block;
      width: 100%;
      border: 0;
      background: transparent;
      padding: 8px 10px;
      border-radius: 6px;
      text-align: left;
      cursor: pointer;
      color: var(--ink);
      font: inherit;
      font-size: 13px;
    }
    .menu button:hover, .menu button:focus-visible {
      background: var(--soft-blue);
      outline: none;
    }
    .menu button:focus-visible {
      outline: 2px solid var(--brand);
    }
    .table-scroll {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 16px;
      background: var(--surface);
    }
    table { width: 100%; border-collapse: collapse; font-size: 13px; }
    th, td { padding: 10px; border-bottom: 1px solid var(--line); text-align: left; white-space: nowrap; }
    th { color: var(--muted); font-weight: 500; background: var(--table-head); }
    td.num { text-align: right; font-variant-numeric: tabular-nums; }
    tbody tr:last-child td { border-bottom: 0; }
    tbody tr:hover td { background: var(--table-hover); }
    .modal-backdrop {
      position: fixed; inset: 0;
      display: none;
      align-items: center; justify-content: center;
      padding: 24px;
      background: var(--modal-backdrop);
      z-index: 50;
    }
    .modal-backdrop.open { display: flex; }
    .modal {
      width: min(860px, 100%);
      max-height: min(780px, 92vh);
      overflow: auto;
      border-radius: 16px;
      background: var(--modal-bg);
      border: 1px solid var(--line-strong);
      box-shadow: 0 18px 48px rgba(55, 53, 47, 0.18);
    }
    .modal-head {
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 16px;
      padding: 18px 20px 14px;
      border-bottom: 1px solid var(--line);
    }
    .modal-head h3 { margin: 0; font-size: 16px; line-height: 1.4; font-weight: 600; }
    .modal-subtitle { margin: 4px 0 0; color: var(--muted); font-size: 14px; line-height: 1.45; }
    .modal-body { padding: 18px 20px 20px; }
    .source-section + .source-section { margin-top: 16px; }
    .source-section h4 { margin: 0 0 8px; color: var(--ink); font-size: 14px; line-height: 1.4; font-weight: 600; }
    .code-wrap { position: relative; }
    pre {
      margin: 0;
      padding: 14px;
      overflow: auto;
      border-radius: 8px;
      border: 1px solid var(--line);
      background: var(--soft);
      color: var(--ink);
      font-size: 12px;
      line-height: 1.5;
    }
    .close {
      width: 32px; height: 32px;
      display: inline-flex; align-items: center; justify-content: center;
      border: 0; border-radius: 999px;
      background: transparent;
      color: var(--muted);
      cursor: pointer; padding: 0;
    }
    .close svg { width: 18px; height: 18px; stroke-width: 2.1; }
    .close:hover, .close:focus-visible { background: var(--soft); outline: none; }
    .close:focus-visible { outline: 2px solid var(--brand); }
    .copy-button {
      position: absolute;
      right: 8px; top: 8px;
      width: 28px; height: 28px;
      display: inline-flex; align-items: center; justify-content: center;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--control-bg);
      color: var(--muted);
      cursor: pointer;
    }
    .copy-button svg { width: 15px; height: 15px; stroke-width: 2; }
    .copy-button:hover, .copy-button:focus-visible { background: var(--soft); color: var(--ink); outline: none; }
    .copy-button:focus-visible { outline: 2px solid var(--brand); }
    @media (max-width: 1100px) {
      .kpi-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    @media (max-width: 900px) {
      .topbar-inner { grid-template-columns: 1fr; }
      .controls { justify-content: flex-start; }
      .kpi-grid { grid-template-columns: 1fr; }
      .dashboard-panel, .dashboard-note, [data-span] { grid-column: 1 / -1; }
    }
    @media (max-width: 620px) {
      .topbar-inner, .dashboard-shell { padding-left: 14px; padding-right: 14px; }
      .segmented { width: 100%; }
      .segmented button { flex: 1; min-width: 0; }
      .date-fields { width: 100%; }
      input[type="date"] { min-width: 0; width: 100%; }
      .chart { height: 240px; min-height: 240px; }
    }
    """

    chart_js = f"""
    const dashboardPayload = {json_script(payload)};
    function cssToken(name) {{
      return getComputedStyle(document.documentElement).getPropertyValue(name).trim();
    }}
    function chartTheme() {{
      return {{
        text: cssToken("--chart-text"),
        muted: cssToken("--chart-muted"),
        line: cssToken("--chart-line"),
        primary: cssToken("--chart-primary"),
        secondary: cssToken("--chart-secondary"),
        tertiary: cssToken("--chart-tertiary"),
        quaternary: cssToken("--chart-quaternary"),
        palette: [1, 2, 3, 4, 5, 6, 7].map(index => cssToken("--chart-" + index))
      }};
    }}
    function axisStyle(extra) {{
      const theme = chartTheme();
      const base = {{
        axisLabel: {{ color: theme.muted }},
        axisLine: {{ lineStyle: {{ color: theme.line }} }},
        axisTick: {{ lineStyle: {{ color: theme.line }} }},
        splitLine: {{ lineStyle: {{ color: theme.line }} }}
      }};
      const merged = Object.assign({{}}, base, extra || {{}});
      merged.axisLabel = Object.assign({{}}, base.axisLabel, (extra || {{}}).axisLabel || {{}});
      return merged;
    }}
    function chartBase(...colorKeys) {{
      const theme = chartTheme();
      return {{
        textStyle: {{ color: theme.text }},
        color: colorKeys.map(key => theme[key] || key)
      }};
    }}
    const categoryColorKeys = {{
      "Amazon": "primary",
      "TikTok Shop": "quaternary",
      "Trustpilot": "tertiary",
      "Shopify Reviews": "secondary",
      "Walmart": "tertiary",
      "TikTok Shop Chat": "quaternary",
      "Zendesk Email": "primary",
      "Instagram DM": "secondary",
      "Amazon Buyer-Seller": "primary",
      "Shopify Inbox": "secondary",
      "Shopify": "secondary",
      "Other": "#8A94A3",
      "Unknown": "#A7ADB6"
    }};
    function categoricalColor(name, index) {{
      const key = String(name || "").trim();
      const theme = chartTheme();
      const tokenOrColor = categoryColorKeys[key];
      return theme[tokenOrColor] || tokenOrColor || theme.palette[index % theme.palette.length];
    }}
    const chartFactories = {{
      reviewTrend: function(type, filteredRows) {{
        const rows = filteredRows("daily");
        const theme = chartTheme();
        return {{
          ...chartBase("primary", "tertiary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: theme.muted }} }},
          grid: {{ left: 52, right: 52, top: 36, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.date), axisLabel: {{ hideOverlap: true }} }}),
          yAxis: [
            axisStyle({{ type: "value", name: "Reviews", axisLabel: {{ formatter: v => Math.round(v) }} }}),
            axisStyle({{ type: "value", name: "Avg ★", min: 1, max: 5, splitLine: {{ show: false }}, axisLabel: {{ formatter: v => Number(v).toFixed(1) }} }})
          ],
          series: [
            {{
              name: "Reviews",
              type: type,
              smooth: type === "line",
              barMaxWidth: 18,
              data: rows.map(row => row.reviews),
              areaStyle: type === "line" ? {{ opacity: 0.08 }} : undefined,
              itemStyle: {{ color: theme.primary }}
            }},
            {{
              name: "Avg ★",
              type: "line",
              yAxisIndex: 1,
              smooth: true,
              data: rows.map(row => row.avg_rating ? Number(row.avg_rating.toFixed(2)) : null),
              itemStyle: {{ color: theme.tertiary }},
              lineStyle: {{ width: 2, type: "dashed" }}
            }}
          ]
        }};
      }},
      ratingMix: function(type, filteredRows) {{
        const counts = filteredRows("ratings").reduce((acc, row) => (acc[row.rating] = row.count, acc), {{}});
        const rows = [1, 2, 3, 4, 5].map(rating => ({{ rating: rating + "★", count: counts[rating] || 0 }}));
        const theme = chartTheme();
        if (type === "pie") {{
          return {{
            ...chartBase(),
            color: rows.map((row, i) => categoricalColor(row.rating, i)),
            tooltip: {{ trigger: "item", valueFormatter: v => Math.round(v) + " reviews" }},
            legend: {{ textStyle: {{ color: theme.muted }} }},
            series: [{{
              type: "pie",
              radius: ["42%", "70%"],
              data: rows.map(row => ({{ name: row.rating, value: row.count }}))
            }}]
          }};
        }}
        return {{
          ...chartBase("primary"),
          tooltip: {{ trigger: "axis" }},
          grid: {{ left: 52, right: 18, top: 28, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.rating) }}),
          yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: v => Math.round(v) }} }}),
          series: [{{ type: "bar", data: rows.map(row => row.count), barMaxWidth: 40, itemStyle: {{ color: theme.primary }} }}]
        }};
      }},
      platformVolume: function(type, filteredRows) {{
        const rows = filteredRows("platforms");
        const labels = rows.map(row => row.segment);
        const theme = chartTheme();
        if (type === "line") {{
          return {{
            ...chartBase("primary", "secondary", "tertiary"),
            tooltip: {{ trigger: "axis" }},
            legend: {{ textStyle: {{ color: theme.muted }} }},
            grid: {{ left: 52, right: 18, top: 36, bottom: 36 }},
            xAxis: axisStyle({{ type: "category", data: labels, axisLabel: {{ rotate: labels.length > 6 ? 28 : 0, hideOverlap: true }} }}),
            yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: v => Math.round(v) }} }}),
            series: [
              {{ name: "Reviews", type: "line", smooth: true, data: rows.map(row => row.reviews), itemStyle: {{ color: theme.primary }} }},
              {{ name: "Tickets", type: "line", smooth: true, data: rows.map(row => row.tickets), itemStyle: {{ color: theme.secondary }} }},
              {{ name: "Q&A", type: "line", smooth: true, data: rows.map(row => row.qa), itemStyle: {{ color: theme.tertiary }} }}
            ]
          }};
        }}
        return {{
          ...chartBase("primary", "secondary", "tertiary"),
          tooltip: {{ trigger: "axis" }},
          legend: {{ textStyle: {{ color: theme.muted }} }},
          grid: {{ left: 52, right: 18, top: 36, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: labels, axisLabel: {{ rotate: labels.length > 6 ? 28 : 0, hideOverlap: true }} }}),
          yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: v => Math.round(v) }} }}),
          series: [
            {{ name: "Reviews", type: "bar", barMaxWidth: 12, stack: "total", data: rows.map(row => row.reviews), itemStyle: {{ color: theme.primary }} }},
            {{ name: "Tickets", type: "bar", barMaxWidth: 12, stack: "total", data: rows.map(row => row.tickets), itemStyle: {{ color: theme.secondary }} }},
            {{ name: "Q&A", type: "bar", barMaxWidth: 12, stack: "total", data: rows.map(row => row.qa), itemStyle: {{ color: theme.tertiary }} }}
          ]
        }};
      }},
      ticketCategories: function(type, filteredRows) {{
        const rows = filteredRows("ticketCategories").slice().sort((a, b) => b.count - a.count);
        const theme = chartTheme();
        if (type === "pie") {{
          return {{
            ...chartBase(),
            color: rows.map((row, i) => categoricalColor(row.category, i)),
            tooltip: {{ trigger: "item", valueFormatter: v => Math.round(v) + " tickets" }},
            legend: {{ type: "scroll", textStyle: {{ color: theme.muted }} }},
            series: [{{ type: "pie", radius: ["42%", "70%"], data: rows.map(row => ({{ name: row.category, value: row.count }})) }}]
          }};
        }}
        return {{
          ...chartBase("quaternary"),
          tooltip: {{ trigger: "axis", valueFormatter: v => Math.round(v) + " tickets" }},
          grid: {{ left: 110, right: 18, top: 18, bottom: 30 }},
          xAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: v => Math.round(v) }} }}),
          yAxis: axisStyle({{ type: "category", data: rows.map(row => row.category) }}),
          series: [{{ type: "bar", data: rows.map(row => row.count), barMaxWidth: 30, itemStyle: {{ color: theme.quaternary }} }}]
        }};
      }},
      qaFunnel: function(type, filteredRows) {{
        const rows = filteredRows("qaFunnel");
        const theme = chartTheme();
        return {{
          ...chartBase("tertiary"),
          tooltip: {{ trigger: "axis", valueFormatter: v => Math.round(v) + " questions" }},
          grid: {{ left: 52, right: 18, top: 28, bottom: 36 }},
          xAxis: axisStyle({{ type: "category", data: rows.map(row => row.stage) }}),
          yAxis: axisStyle({{ type: "value", axisLabel: {{ formatter: v => Math.round(v) }} }}),
          series: [
            {{
              type: type === "line" ? "line" : "bar",
              data: rows.map(row => row.count),
              smooth: type === "line",
              barMaxWidth: 48,
              itemStyle: {{ color: theme.tertiary }},
              areaStyle: type === "line" ? {{ opacity: 0.12 }} : undefined
            }}
          ]
        }};
      }}
    }};
    const sourceMap = {json_script(source_map)};
    setupDashboardRuntime({{
      datasets: dashboardPayload.datasets,
      availableDates: dashboardPayload.availableDates,
      defaultRange: dashboardPayload.defaultRange,
      initialCharts: {json_script(initial_charts)},
      chartFactories,
      sourceMap,
      tables: {json_script(table_config)},
      fullScript: {js_string(ANALYSIS_LOGIC)},
      modalSubtitlePrefix: "Panel transform for "
    }});
    """

    return f"""<!-- Generated by Trae Work -->
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(payload["title"])}</title>
  <style>{css}</style>
</head>
<body>
  <header class="topbar">
    <div class="topbar-inner">
      <div>
        <h1>{html.escape(payload["title"])}</h1>
        <p class="subtitle">{html.escape(payload["subtitle"])}</p>
        <p class="freshness" id="dataFreshness">Latest data: {html.escape(payload["freshness"]["latestDataDate"])} | Earliest: {html.escape(payload["freshness"]["earliestDataDate"])} | Captured: {html.escape(payload["freshness"]["latestCapturedAt"])} | {html.escape(payload["timezone"])}</p>
      </div>
      <div class="controls" aria-label="Dashboard time controls">
        <span class="range-label" id="activeRangeLabel"></span>
        <div class="segmented" aria-label="Time preset">
          <button data-range-preset="7D">7D</button>
          <button data-range-preset="30D">30D</button>
          <button data-range-preset="MTD">MTD</button>
          <button data-range-preset="QTD">QTD</button>
          <button data-range-preset="YTD">YTD</button>
          <button data-range-preset="ALL">All</button>
        </div>
        <div class="date-fields">
          <input id="rangeStart" data-range-input type="date" aria-label="Start date">
          <input id="rangeEnd" data-range-input type="date" aria-label="End date">
        </div>
        <div class="segmented theme-switch" aria-label="Theme">
          <button data-theme-choice="light" type="button" aria-label="Light theme" title="Light">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
              <circle cx="12" cy="12" r="4"></circle>
              <path d="M12 2v2"></path>
              <path d="M12 20v2"></path>
              <path d="m4.93 4.93 1.41 1.41"></path>
              <path d="m17.66 17.66 1.41 1.41"></path>
              <path d="M2 12h2"></path>
              <path d="M20 12h2"></path>
              <path d="m6.34 17.66-1.41 1.41"></path>
              <path d="m19.07 4.93-1.41 1.41"></path>
            </svg>
          </button>
          <button data-theme-choice="trae-dark" type="button" aria-label="Dark theme" title="Dark">
            <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
              <path d="M20.99 13.53A8.5 8.5 0 1 1 10.47 3.01 7 7 0 0 0 20.99 13.53Z"></path>
            </svg>
          </button>
        </div>
      </div>
    </div>
  </header>
  <main class="dashboard-shell">
    {content}
  </main>
  <div id="modalBackdrop" class="modal-backdrop" role="dialog" aria-modal="true">
    <section class="modal">
      <div class="modal-head">
        <div>
          <h3 id="modalTitle">Data Source</h3>
          <p class="modal-subtitle" id="modalSubtitle"></p>
        </div>
        <button class="close" aria-label="Close" onclick="closeModal()">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
            <path d="M18 6 6 18"></path>
            <path d="m6 6 12 12"></path>
          </svg>
        </button>
      </div>
      <div class="modal-body">
        <section class="source-section">
          <h4>Panel transform</h4>
          <div class="code-wrap">
            <button class="copy-button" aria-label="Copy panel transform" onclick="copyCode('modalSnippet', this)">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
                <rect x="9" y="9" width="11" height="11" rx="2"></rect>
                <path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path>
              </svg>
            </button>
            <pre><code id="modalSnippet"></code></pre>
          </div>
        </section>
        <section class="source-section">
          <h4>Analysis logic</h4>
          <div class="code-wrap">
            <button class="copy-button" aria-label="Copy analysis logic" onclick="copyCode('modalCode', this)">
              <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" aria-hidden="true">
                <rect x="9" y="9" width="11" height="11" rx="2"></rect>
                <path d="M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1"></path>
              </svg>
            </button>
            <pre><code id="modalCode"></code></pre>
          </div>
        </section>
      </div>
    </section>
  </div>
  <script>{echarts}</script>
  <script>{runtime}</script>
  <script>{chart_js}</script>
</body>
</html>
"""


def main() -> None:
    rows = normalize_snapshots(read_sources())
    payload = make_dashboard_payload(rows)
    DASHBOARD_DATA.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    DASHBOARD_HTML.write_text(build_html(payload), encoding="utf-8")
    print(f"Wrote {DASHBOARD_HTML}")
    print(f"Wrote {DASHBOARD_DATA}")
    print(f"Total events: {len(rows)}")
    print(f"Date range:  {payload['freshness']['earliestDataDate']} -> {payload['freshness']['latestDataDate']}")


if __name__ == "__main__":
    main()
